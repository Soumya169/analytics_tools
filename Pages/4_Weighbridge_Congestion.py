import streamlit as st
import pandas as pd
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from utils import load_file, to_dt, auto_index, sec_to_hms, min_to_hms
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Weighbridge Congestion", page_icon="⚖️", layout="wide")

st.title("⚖️ Weighbridge Congestion Analysis")
st.markdown("Analyse weighbridge utilisation, waiting time, and congestion by bridge / shift / hour.")
st.markdown("---")

# ── INBOUND / OUTBOUND TOGGLE ────────────────────────────────
col_btn1, col_btn2, _ = st.columns([1, 1, 4])
with col_btn1:
    ib = st.button("📥 Inbound", use_container_width=True,
                   type="primary" if st.session_state.get("wb_mode","inbound") == "inbound" else "secondary")
with col_btn2:
    ob = st.button("📤 Outbound", use_container_width=True,
                   type="primary" if st.session_state.get("wb_mode","inbound") == "outbound" else "secondary")
if ib: st.session_state["wb_mode"] = "inbound"
if ob: st.session_state["wb_mode"] = "outbound"
mode = st.session_state.get("wb_mode", "inbound")

st.markdown(f"**Mode: {'📥 Inbound' if mode == 'inbound' else '📤 Outbound'}**")

if mode == "inbound":
    st.markdown("**Inbound Weighbridge Stages:** `GateIn → GrossWeight (GI-GW)` and `GrossWeight → TareWeight (GW-TW)`")
else:
    st.markdown("**Outbound Weighbridge Stages:** `GrossWeight → LoadingIn (GW-LI)` and `LoadingOut → TareWeight (LO-TW)`")

st.markdown("---")

uploaded = st.file_uploader("📂 Upload Excel File", type=["xlsx","xls"], key="wb_upload")
if uploaded is None:
    st.info("👆 Upload your Excel file to begin Weighbridge Congestion Analysis.")
    st.markdown("""
    **Expected columns:**
    - `First Weighbridge No` / `Second WeighBridge No` — weighbridge identifier
    - `GrossWeight` — datetime of gross weighment
    - `TareWeight` — datetime of tare weighment
    - `GateIn` — gate in datetime (inbound)
    - `Shift` — shift (A/B/C)
    - `Transporter Name` — transporter
    """)
    st.stop()

df = load_file(uploaded)
total_rows = len(df)
st.success(f"✅ Loaded: **{total_rows} rows, {len(df.columns)} columns**")

with st.expander("🔍 Debug Info"):
    st.write("**Columns:**", df.columns.tolist())
    st.dataframe(df.head(3), use_container_width=True)

st.markdown("---")
st.subheader("🔧 Map Your Columns")
all_cols = ["-- Not Available --"] + df.columns.tolist()

c1, c2, c3 = st.columns(3)
with c1:
    col_wb1     = st.selectbox("First Weighbridge No",   all_cols, index=auto_index(all_cols,"First Weighbridge No"))
    col_wb2     = st.selectbox("Second Weighbridge No",  all_cols, index=auto_index(all_cols,"Second WeighBridge No"))
with c2:
    col_grosswt = st.selectbox("GrossWeight (datetime)", all_cols, index=auto_index(all_cols,"GrossWeight"))
    col_tarewt  = st.selectbox("TareWeight (datetime)",  all_cols, index=auto_index(all_cols,"TareWeight"))
with c3:
    col_gatein  = st.selectbox("GateIn (datetime)",      all_cols, index=auto_index(all_cols,"GateIn"))
    col_shift   = st.selectbox("Shift",                  all_cols, index=auto_index(all_cols,"Shift"))

if mode == "outbound":
    st.markdown("**Outbound additional columns:**")
    o1, o2 = st.columns(2)
    with o1:
        col_loadingin  = st.selectbox("LoadingIn (datetime)",  all_cols, index=auto_index(all_cols,"LoadingIn"),  key="wb_li")
    with o2:
        col_loadingout = st.selectbox("LoadingOut (datetime)", all_cols, index=auto_index(all_cols,"LoadingOut"), key="wb_lo")
else:
    col_loadingin  = "-- Not Available --"
    col_loadingout = "-- Not Available --"

st.markdown("---")

if st.button("📊 Run Weighbridge Analysis", type="primary", use_container_width=True):

    result = df.copy()

    dt_gatein  = to_dt(result[col_gatein])  if col_gatein  != "-- Not Available --" else None
    dt_grosswt = to_dt(result[col_grosswt]) if col_grosswt != "-- Not Available --" else None
    dt_tarewt  = to_dt(result[col_tarewt])  if col_tarewt  != "-- Not Available --" else None

    # Extract hour from GrossWeight for congestion by hour
    if dt_grosswt is not None:
        result["GrossWt Hour"] = dt_grosswt.dt.hour
        result["GrossWt Date"] = dt_grosswt.dt.date

    # ── INBOUND CALCULATIONS ──────────────────────────────────
    if mode == "inbound":
        if dt_gatein is not None and dt_grosswt is not None:
            diff = (dt_grosswt - dt_gatein).dt.total_seconds()
            result["GI-GW (Wait at WB)"] = diff.apply(sec_to_hms)
            result["GI-GW_min"]           = diff / 60
            st.success(f"✅ GI-GW (Gate In to Gross Weighment) calculated — {int((diff>=0).sum())} rows")

        if dt_grosswt is not None and dt_tarewt is not None:
            diff2 = (dt_tarewt - dt_grosswt).dt.total_seconds()
            result["GW-TW (Processing)"] = diff2.apply(sec_to_hms)
            result["GW-TW_min"]           = diff2 / 60
            st.success(f"✅ GW-TW (Gross to Tare) calculated — {int((diff2>=0).sum())} rows")

    # ── OUTBOUND CALCULATIONS ─────────────────────────────────
    else:
        dt_loadingin  = to_dt(result[col_loadingin])  if col_loadingin  != "-- Not Available --" else None
        dt_loadingout = to_dt(result[col_loadingout]) if col_loadingout != "-- Not Available --" else None

        if dt_grosswt is not None and dt_loadingin is not None:
            diff = (dt_loadingin - dt_grosswt).dt.total_seconds()
            result["GW-LI (Wait for Loading)"] = diff.apply(sec_to_hms)
            result["GW-LI_min"]                 = diff / 60
            st.success(f"✅ GW-LI (Gross Weighment to Loading Start) — {int((diff>=0).sum())} rows")

        if dt_loadingout is not None and dt_tarewt is not None:
            diff2 = (dt_tarewt - dt_loadingout).dt.total_seconds()
            result["LO-TW (Wait for Tare)"] = diff2.apply(sec_to_hms)
            result["LO-TW_min"]              = diff2 / 60
            st.success(f"✅ LO-TW (Loading Out to Tare Weighment) — {int((diff2>=0).sum())} rows")

    st.markdown("---")

    # ── WEIGHBRIDGE SUMMARY ───────────────────────────────────
    min_cols = [c for c in result.columns if c.endswith("_min")]

    if col_wb1 != "-- Not Available --":
        st.subheader("⚖️ First Weighbridge Performance")
        wb1_grp = result.groupby(col_wb1).size().reset_index(name="Trip Count")
        if min_cols:
            for mc in min_cols:
                lbl = mc.replace("_min","")
                wb1_avg = result.groupby(col_wb1)[mc].mean().round(2).reset_index()
                wb1_avg[f"Avg {lbl} (HH:MM:SS)"] = wb1_avg[mc].apply(min_to_hms)
                wb1_grp = wb1_grp.merge(wb1_avg[[col_wb1, f"Avg {lbl} (HH:MM:SS)"]], on=col_wb1, how="left")
        st.dataframe(wb1_grp, use_container_width=True)

    if col_wb2 != "-- Not Available --":
        st.markdown("---")
        st.subheader("⚖️ Second Weighbridge Performance")
        wb2_grp = result.groupby(col_wb2).size().reset_index(name="Trip Count")
        st.dataframe(wb2_grp, use_container_width=True)

    # Hour-wise congestion
    if "GrossWt Hour" in result.columns:
        st.markdown("---")
        st.subheader("🕐 Hour-wise Congestion (Trip Count by Hour)")
        hour_grp = result.groupby("GrossWt Hour").size().reset_index(name="Trip Count")
        hour_grp["Hour"] = hour_grp["GrossWt Hour"].apply(lambda h: f"{h:02d}:00 - {h:02d}:59")
        st.dataframe(hour_grp[["Hour","Trip Count"]], use_container_width=True)

    # Shift-wise
    if col_shift != "-- Not Available --":
        st.markdown("---")
        st.subheader("🕐 Shift-wise Weighbridge Congestion")
        shift_grp = result.groupby(col_shift).size().reset_index(name="Trip Count")
        if min_cols:
            for mc in min_cols:
                lbl = mc.replace("_min","")
                s_avg = result.groupby(col_shift)[mc].mean().round(2).reset_index()
                s_avg[f"Avg {lbl}"] = s_avg[mc].apply(min_to_hms)
                shift_grp = shift_grp.merge(s_avg[[col_shift, f"Avg {lbl}"]], on=col_shift, how="left")
        st.dataframe(shift_grp, use_container_width=True)

    # Full preview
    st.markdown("---")
    st.subheader(f"👁 Full Data — {len(result)} Rows")
    st.dataframe(result, use_container_width=True, height=400)

    # Download
    st.markdown("---")
    buf = io.BytesIO()

    # Drop helper _min columns from download
    download_df = result.drop(columns=[c for c in result.columns if c.endswith("_min")], errors='ignore')

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        download_df.to_excel(writer, index=False, sheet_name="WB Data")
        if col_wb1 != "-- Not Available --":
            wb1_grp.to_excel(writer, index=False, sheet_name="WB1 Summary")
        if "hour_grp" in dir():
            hour_grp.to_excel(writer, index=False, sheet_name="Hour Congestion")
        if col_shift != "-- Not Available --":
            shift_grp.to_excel(writer, index=False, sheet_name="Shift Summary")

        for sname in writer.sheets:
            ws = writer.sheets[sname]
            for cell in ws[1]:
                cell.fill = PatternFill("solid", start_color="833C00")
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 22
            ws.freeze_panes = "A2"

    buf.seek(0)
    st.download_button(
        f"⬇️ Download Weighbridge Analysis Excel ({len(download_df)} rows)",
        data=buf,
        file_name="Weighbridge_Congestion_Analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, type="primary"
    )