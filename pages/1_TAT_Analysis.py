import streamlit as st
import pandas as pd
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from utils import load_file, to_dt, auto_index, parse_summary, calculate_stages, hms_to_min, min_to_hms, hms_to_min_series, hms_to_excel_fraction_series
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="TAT Analysis", page_icon="📥", layout="wide")

st.set_page_config(page_title="TAT Analysis", page_icon="📥", layout="wide")

# ── TOP RIGHT TEMPLATE BUTTONS ─────────────────────────────
_, col_ib, col_ob = st.columns([6, 1, 1])
with col_ib:
    ib_headers = ["Trip ID","Vehicle Number","Transporter Name","Shift",
                  "Gate Entry Type","Supplier Name","Mat. Group","YardIn",
                  "GateIn","GrossWeight","TareWeight","GateOut","Net Weight","Remarks"]
    ib_buf = io.BytesIO()
    with pd.ExcelWriter(ib_buf, engine='xlsxwriter') as w:
        pd.DataFrame(columns=ib_headers).to_excel(w, index=False, sheet_name="IB Template")
        fmt = w.book.add_format({'bold':True,'font_color':'#FFFFFF','bg_color':'#1F4E79','align':'center','border':1})
        ws  = w.sheets["IB Template"]
        for i,h in enumerate(ib_headers):
            ws.write(0,i,h,fmt)
            ws.set_column(i,i,18)
    ib_buf.seek(0)
    st.download_button("📥 IB Template", data=ib_buf,
        file_name="IB_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="ib_tmpl_top", use_container_width=True)

with col_ob:
    ob_headers = ["Trip ID","Vehicle Number","Transporter Name","Shift",
                  "Gate Entry Type","Supplier Name","Mat. Group","YardIn","ParkIn",
                  "YardOut","ParkOut","GateIn","GrossWeight","LoadingIn","LoadingOut",
                  "TareWeight","GateOut","Unloader Alias","Net Weight","Remarks"]
    ob_buf = io.BytesIO()
    with pd.ExcelWriter(ob_buf, engine='xlsxwriter') as w:
        pd.DataFrame(columns=ob_headers).to_excel(w, index=False, sheet_name="OB Template")
        fmt = w.book.add_format({'bold':True,'font_color':'#FFFFFF','bg_color':'#833C00','align':'center','border':1})
        ws  = w.sheets["OB Template"]
        for i,h in enumerate(ob_headers):
            ws.write(0,i,h,fmt)
            ws.set_column(i,i,18)
    ob_buf.seek(0)
    st.download_button("📤 OB Template", data=ob_buf,
        file_name="OB_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="ob_tmpl_top", use_container_width=True)
st.title("📥📤 TAT Analysis")
st.markdown("Calculate TAT for **Inbound** and **Outbound** trips in `HH:MM:SS` format.")
st.markdown("---")

# ── INBOUND / OUTBOUND TOGGLE ─────────────────────────────────
col_btn1, col_btn2, _ = st.columns([1, 1, 4])
with col_btn1:
    if st.button("📥 Inbound", use_container_width=True,
                 type="primary" if st.session_state.get("tat_mode","inbound") == "inbound" else "secondary"):
        st.session_state["tat_mode"] = "inbound"
        for k in ["tat_result","tat_stats","tat_groupby","tat_time_data",
                  "tat_dt_cols","tat_tat_set","tat_total","tat_gateout_col"]:
            st.session_state.pop(k, None)
        st.rerun()
with col_btn2:
    if st.button("📤 Outbound", use_container_width=True,
                 type="primary" if st.session_state.get("tat_mode","inbound") == "outbound" else "secondary"):
        st.session_state["tat_mode"] = "outbound"
        for k in ["tat_result","tat_stats","tat_groupby","tat_time_data",
                  "tat_dt_cols","tat_tat_set","tat_total","tat_gateout_col"]:
            st.session_state.pop(k, None)
        st.rerun()

mode = st.session_state.get("tat_mode", "inbound")
st.markdown("---")

# ─────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────
WEEKDAY_ORDER = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
MONTH_ORDER   = ["January","February","March","April","May","June",
                 "July","August","September","October","November","December"]

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def hms_str_to_fraction(hms_str):
    try:
        if not hms_str or hms_str == "–" or pd.isna(hms_str): return None
        parts = str(hms_str).strip().split(":")
        if len(parts) != 3: return None
        h,m,s = int(parts[0]),int(parts[1]),int(parts[2])
        return (h*3600 + m*60 + s) / 86400
    except: return None


def build_stats_table(result, tat_cols):
    rows = []
    for col in tat_cols:
        if col not in result.columns: continue
        mins = hms_to_min_series(result[col]).dropna()   # vectorized
        if len(mins) == 0: continue
        rows.append({
            "TAT Stage":  col,
            "Average":    min_to_hms(mins.mean()),
            "Median":     min_to_hms(mins.median()),
            "Min":        min_to_hms(mins.min()),
            "Max":        min_to_hms(mins.max()),
            "Valid Rows": int(mins.count()),
            "Blank Rows": int(len(result) - mins.count()),
        })
    return pd.DataFrame(rows)


def build_groupby_stats(result, tat_cols, group_col):
    if not group_col or group_col not in result.columns: return pd.DataFrame()
    rows = []
    tmp = result.copy()
    for col in tat_cols:
        if col in tmp.columns:
            tmp[col+"_min"] = hms_to_min_series(tmp[col])   # vectorized
    for col in tat_cols:
        mc = col+"_min"
        if mc not in tmp.columns: continue
        grp = tmp.groupby(group_col)[mc]
        avg,med,cnt = grp.mean(),grp.median(),grp.count()
        for cat in avg.index:
            rows.append({"Category":str(cat),"TAT Stage":col,
                         "Average":min_to_hms(avg[cat]),"Median":min_to_hms(med[cat]),
                         "Trip Count":int(cnt[cat])})
    return pd.DataFrame(rows)


def build_time_dimension(result, tat_cols, group_col, sort_order=None):
    """Generic: group by any time column, compute Avg/Med/Min/Max/Count per TAT stage."""
    if group_col not in result.columns: return pd.DataFrame()
    rows = []
    tmp = result.copy()
    for col in tat_cols:
        if col in tmp.columns:
            tmp[col+"_min"] = hms_to_min_series(tmp[col])   # vectorized
    for col in tat_cols:
        mc = col+"_min"
        if mc not in tmp.columns: continue
        grp  = tmp.groupby(group_col)[mc]
        avg  = grp.mean(); med = grp.median()
        mn   = grp.min();  mx  = grp.max(); cnt = grp.count()
        keys = [k for k in (sort_order or []) if k in avg.index]
        keys += [k for k in avg.index if k not in keys]
        for k in keys:
            rows.append({group_col:str(k),"TAT Stage":col,"Trip Count":int(cnt[k]),
                         "Average":min_to_hms(avg[k]),"Median":min_to_hms(med[k]),
                         "Min":min_to_hms(mn[k]),"Max":min_to_hms(mx[k])})
    return pd.DataFrame(rows)


def make_pivot(df, index_col, tat_cols, metric="Average"):
    """Create Excel-style pivot: rows = index_col values, columns = TAT stages."""
    if df.empty: return pd.DataFrame()
    rows = []
    unique_keys = df[index_col].unique().tolist()
    for key in unique_keys:
        row = {index_col: key}
        sub = df[df[index_col]==key]
        trips = 0
        for _, r in sub.iterrows():
            row[r["TAT Stage"]] = r[metric]
            trips = r["Trip Count"]
        row["Trip Count"] = trips
        rows.append(row)
    return pd.DataFrame(rows)


def show_stat_cards(stats_df):
    if stats_df.empty: return
    cols = st.columns(len(stats_df))
    for i, row in stats_df.iterrows():
        with cols[i]:
            st.markdown(f"""
            <div style='background:#1F4E79;padding:12px;border-radius:10px;
                        text-align:center;margin-bottom:6px'>
                <div style='color:#BDD7EE;font-size:10px;text-transform:uppercase;
                            letter-spacing:1px'>{row['TAT Stage']}</div>
                <div style='color:#00d4ff;font-size:18px;font-weight:700;
                            margin:4px 0'>{row['Average']}</div>
                <div style='color:#94a3b8;font-size:10px'>Average</div>
            </div>
            <div style='background:#375623;padding:12px;border-radius:10px;text-align:center'>
                <div style='color:#E2EFDA;font-size:10px;text-transform:uppercase;
                            letter-spacing:1px'>{row['TAT Stage']}</div>
                <div style='color:#10b981;font-size:18px;font-weight:700;
                            margin:4px 0'>{row['Median']}</div>
                <div style='color:#94a3b8;font-size:10px'>Median</div>
            </div>""", unsafe_allow_html=True)


def render_view_table(df, view, extra_cols=None):
    if df.empty: st.info("No data."); return
    base = extra_cols or []
    if   view == "📈 Average": cols = base + ["Average","Trip Count"]
    elif view == "📉 Median":  cols = base + ["Median","Trip Count"]
    else:                       cols = base + ["Average","Median","Min","Max","Trip Count"]
    st.dataframe(df[[c for c in cols if c in df.columns]],
                 use_container_width=True, hide_index=True)


def render_cards(df, label_col, stage_filter=None):
    """Render visual cards for any time dimension."""
    subset = df[df["TAT Stage"]==stage_filter] if stage_filter else df
    if subset.empty: return
    n = min(len(subset), 7)
    day_cols = st.columns(n)
    for i,(_, row) in enumerate(subset.iterrows()):
        if i >= n: break
        with day_cols[i]:
            st.markdown(f"""
            <div style='background:#833C00;padding:8px;border-radius:8px;
                        text-align:center;margin-bottom:4px'>
                <div style='color:#FCE4D6;font-size:9px;font-weight:700'>{row[label_col]}</div>
            </div>
            <div style='background:#1F4E79;padding:7px;border-radius:7px;
                        text-align:center;margin-bottom:3px'>
                <div style='color:#BDD7EE;font-size:9px'>Avg</div>
                <div style='color:#00d4ff;font-size:12px;font-weight:700'>{row['Average']}</div>
            </div>
            <div style='background:#375623;padding:7px;border-radius:7px;
                        text-align:center;margin-bottom:3px'>
                <div style='color:#E2EFDA;font-size:9px'>Median</div>
                <div style='color:#10b981;font-size:12px;font-weight:700'>{row['Median']}</div>
            </div>
            <div style='background:#243F60;padding:5px;border-radius:7px;text-align:center'>
                <div style='color:#BDD7EE;font-size:9px'>Trips</div>
                <div style='color:white;font-size:12px;font-weight:700'>{row['Trip Count']}</div>
            </div>""", unsafe_allow_html=True)


def format_stats_sheet(ws, max_row):
    STATS_TIME_COLS = {"Average","Median","Min","Max"}
    headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    for col_idx, header in enumerate(headers, 1):
        if header in STATS_TIME_COLS:
            for rn in range(2, max_row+1):
                cell = ws.cell(rn, col_idx)
                frac = hms_str_to_fraction(cell.value)
                if frac is not None:
                    cell.value = frac
                    cell.number_format = "[HH]:MM:SS"
                cell.alignment = Alignment(horizontal="center", vertical="center")


def style_sheet(ws, header_color):
    thin = Border(left=Side(style='thin'),right=Side(style='thin'),
                  top=Side(style='thin'),bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.fill = PatternFill("solid", start_color=header_color)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 28
    for row_num in range(2, ws.max_row+1):
        bg = "EBF3FB" if row_num % 2 == 0 else "FFFFFF"
        for cell in ws[row_num]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", start_color=bg)
            cell.border = thin
    for col_idx in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.freeze_panes = "A2"


def export_excel(result, stats_df, groupby_df, time_data, dt_col_names, tat_cols_set):
    # ── Pre-convert before writing (avoid cell-by-cell loops) ──
    result_out = result.copy()
    for col in tat_cols_set:
        if col in result_out.columns:
            result_out[col] = hms_to_excel_fraction_series(result_out[col])
    for col in dt_col_names:
        if col in result_out.columns:
            result_out[col] = pd.to_datetime(
                result_out[col].replace("", pd.NaT), dayfirst=True, errors='coerce'
            ).dt.to_pydatetime()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # ── Write all sheets ──────────────────────────────────
        result_out.to_excel(writer, index=False, sheet_name="Full Data")
        if not stats_df.empty:
            stats_df.to_excel(writer, index=False, sheet_name="Overall Stats")
        if not groupby_df.empty:
            groupby_df.to_excel(writer, index=False, sheet_name="Category Stats")

        # Time dimension pivot sheets
        sheet_specs = [
            ("date_df",    "Date Wise",    "1F4E79", "GateOut Date"),
            ("dayofwk_df", "Day of Week",  "375623", "GateOut DayOfWeek"),
            ("week_df",    "Week Wise",    "833C00", "GateOut WeekNo"),
            ("month_df",   "Month Wise",   "7B2D8B", "GateOut Month"),
        ]
        for key, sheet_name, color, col in sheet_specs:
            df_td = time_data.get(key, pd.DataFrame())
            if not df_td.empty:
                df_td.to_excel(writer, index=False, sheet_name=sheet_name)

        # ── Style all sheets ──────────────────────────────────
        sheet_colors = {
            "Full Data":     "1F4E79",
            "Overall Stats": "375623",
            "Category Stats":"7B2D8B",
            "Date Wise":     "1F4E79",
            "Day of Week":   "375623",
            "Week Wise":     "833C00",
            "Month Wise":    "7B2D8B",
        }
        for sname, ws in writer.sheets.items():
            style_sheet(ws, sheet_colors.get(sname, "1F4E79"))

        # ── Fix Full Data sheet formats ───────────────────────
        ws_full = writer.sheets["Full Data"]
        col_letter_map = {col: get_column_letter(idx)
                          for idx, col in enumerate(result_out.columns, 1)}

        # Datetime columns → already converted, just set format
        for col_name in dt_col_names:
            if col_name in col_letter_map:
                cl = col_letter_map[col_name]
                for rn in range(2, len(result_out)+2):
                    cell = ws_full[f"{cl}{rn}"]
                    if cell.value and not isinstance(cell.value, str):
                        cell.number_format = "DD-MM-YYYY HH:MM:SS"

        # TAT columns → already fractions, just apply format + style
        for col_name in tat_cols_set:
            if col_name in col_letter_map:
                col_idx = result_out.columns.tolist().index(col_name)+1
                ws_full.cell(1, col_idx).fill = PatternFill("solid", start_color="375623")
                ws_full.cell(1, col_idx).font = Font(name="Arial",size=10,bold=True,color="FFFFFF")
                cl = get_column_letter(col_idx)
                for rn in range(2, len(result_out)+2):
                    cell = ws_full[f"{cl}{rn}"]
                    if cell.value not in (None,""):
                        cell.number_format = "[HH]:MM:SS"
                    cell.fill = PatternFill("solid", start_color="E2EFDA")
                    cell.font = Font(name="Arial",size=9,bold=True,color="375623")

        # Time dimension columns in Full Data → highlight each differently
        time_col_styles = {
            "GateOut Date":      ("FCE4D6","833C00","DD-MM-YYYY"),
            "GateOut DayOfWeek": ("E2EFDA","375623","@"),
            "GateOut WeekNo":    ("FFF2CC","7F6000","@"),
            "GateOut Month":     ("EAD1DC","7B2D8B","@"),
        }
        for col_name,(bg,fg,nfmt) in time_col_styles.items():
            if col_name in result.columns:
                col_idx = result.columns.tolist().index(col_name)+1
                cl = get_column_letter(col_idx)
                ws_full.cell(1,col_idx).fill = PatternFill("solid",start_color=fg)
                ws_full.cell(1,col_idx).font = Font(name="Arial",size=10,bold=True,color="FFFFFF")
                for rn in range(2, len(result_out)+2):
                    cell = ws_full[f"{cl}{rn}"]
                    cell.fill = PatternFill("solid",start_color=bg)
                    cell.font = Font(name="Arial",size=9,bold=True,color=fg)
                    cell.number_format = nfmt
                    cell.alignment = Alignment(horizontal="center",vertical="center")

        # ── Fix time dimension stats sheets → [HH]:MM:SS ─────
        for _, sheet_name, _, _ in sheet_specs:
            if sheet_name in writer.sheets:
                td_key = [k for k,s,_,_ in sheet_specs if s==sheet_name][0]
                df_td = time_data.get(td_key, pd.DataFrame())
                if not df_td.empty:
                    format_stats_sheet(writer.sheets[sheet_name], len(df_td)+1)

        # Fix Overall Stats and Category Stats
        if "Overall Stats"  in writer.sheets and not stats_df.empty:
            format_stats_sheet(writer.sheets["Overall Stats"],  len(stats_df)+1)
        if "Category Stats" in writer.sheets and not groupby_df.empty:
            format_stats_sheet(writer.sheets["Category Stats"], len(groupby_df)+1)

    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────
# RENDER ANALYSIS (from session_state)
# ─────────────────────────────────────────────────────────────
def render_analysis():
    result     = st.session_state["tat_result"]
    stats_df   = st.session_state["tat_stats"]
    groupby_df = st.session_state["tat_groupby"]
    time_data  = st.session_state["tat_time_data"]
    dt_cols    = st.session_state["tat_dt_cols"]
    tat_set    = st.session_state["tat_tat_set"]
    total_rows = st.session_state["tat_total"]
    sel_cat    = st.session_state.get("tat_sel_cat","-- None --")
    tat_cols   = list(tat_set)

    # ── 1. OVERALL STATS ─────────────────────────────────────
    st.markdown("---")
    st.subheader("📊 Overall TAT — Average & Median")
    show_stat_cards(stats_df)
    st.markdown("")
    view = st.radio("Select View",["📈 Average","📉 Median","📊 Both"],
                    horizontal=True, key="tat_view_radio",
                    index=["📈 Average","📉 Median","📊 Both"].index(
                        st.session_state.get("tat_view","📊 Both")))
    st.session_state["tat_view"] = view
    if not stats_df.empty:
        render_view_table(stats_df, view,
                          extra_cols=["TAT Stage","Valid Rows","Blank Rows"])

    # ── 2. CATEGORY GROUP-BY ─────────────────────────────────
    if sel_cat != "-- None --" and not groupby_df.empty:
        st.markdown("---")
        st.subheader(f"📋 TAT by {sel_cat}")
        gb_view = st.radio("Category View",["📈 Average","📉 Median","📊 Both"],
                           horizontal=True, key="tat_gb_view")
        for tat_col in tat_cols:
            subset = groupby_df[groupby_df["TAT Stage"]==tat_col].copy()
            if subset.empty: continue
            st.markdown(f"**{tat_col}**")
            render_view_table(subset, gb_view, extra_cols=["Category"])

    # ── 3. TIME DIMENSION ANALYSIS ───────────────────────────
    st.markdown("---")
    st.subheader("📅 Time Dimension Analysis")
    st.markdown("Analyse TAT across different time groupings from GateOut timestamp.")

    # Tab-based navigation for 4 time dimensions
    tab_date, tab_day, tab_week, tab_month = st.tabs([
        "📆 Date Wise", "📅 Day of Week", "🗓️ Week Wise", "🗃️ Month Wise"
    ])

    # ── TAB 1: DATE WISE ─────────────────────────────────────
    with tab_date:
        df_date = time_data.get("date_df", pd.DataFrame())
        if df_date.empty:
            st.info("No GateOut date data available.")
        else:
            st.markdown("**Grouped by exact GateOut Date (DD-MM-YYYY)**")
            available = df_date["TAT Stage"].unique().tolist()
            sel_stage = st.selectbox("TAT Stage", ["All Stages"]+available, key="td_date_stage")
            dv = st.radio("View", ["📈 Average","📉 Median","📊 Both"],
                          horizontal=True, key="td_date_view")
            st.markdown("")
            if sel_stage == "All Stages":
                st.markdown("##### 📊 Pivot — Dates × TAT Stages")
                metric = "Average" if dv=="📈 Average" else "Median"
                pivot = make_pivot(df_date, "GateOut Date", tat_cols, metric)
                st.dataframe(pivot, use_container_width=True, hide_index=True)
            else:
                sub = df_date[df_date["TAT Stage"]==sel_stage].copy()
                render_view_table(sub, dv, extra_cols=["GateOut Date"])
                render_cards(sub, "GateOut Date", sel_stage)

    # ── TAB 2: DAY OF WEEK ───────────────────────────────────
    with tab_day:
        df_dow = time_data.get("dayofwk_df", pd.DataFrame())
        if df_dow.empty:
            st.info("No GateOut day-of-week data available.")
        else:
            st.markdown("**Grouped by Day of Week (Monday → Sunday)**")
            available = df_dow["TAT Stage"].unique().tolist()
            sel_stage = st.selectbox("TAT Stage", ["All Stages"]+available, key="td_dow_stage")
            dv = st.radio("View", ["📈 Average","📉 Median","📊 Both"],
                          horizontal=True, key="td_dow_view")
            st.markdown("")
            if sel_stage == "All Stages":
                st.markdown("##### 📊 Pivot — Day of Week × TAT Stages")
                metric = "Average" if dv=="📈 Average" else "Median"
                pivot = make_pivot(df_dow, "GateOut DayOfWeek", tat_cols, metric)
                st.dataframe(pivot, use_container_width=True, hide_index=True)
            else:
                sub = df_dow[df_dow["TAT Stage"]==sel_stage].copy()
                render_view_table(sub, dv, extra_cols=["GateOut DayOfWeek"])
                render_cards(sub, "GateOut DayOfWeek", sel_stage)

    # ── TAB 3: WEEK WISE ─────────────────────────────────────
    with tab_week:
        df_week = time_data.get("week_df", pd.DataFrame())
        if df_week.empty:
            st.info("No GateOut week data available.")
        else:
            st.markdown("**Grouped by Week Number (Week 1, Week 2 …)**")
            available = df_week["TAT Stage"].unique().tolist()
            sel_stage = st.selectbox("TAT Stage", ["All Stages"]+available, key="td_week_stage")
            dv = st.radio("View", ["📈 Average","📉 Median","📊 Both"],
                          horizontal=True, key="td_week_view")
            st.markdown("")
            if sel_stage == "All Stages":
                st.markdown("##### 📊 Pivot — Week Number × TAT Stages")
                metric = "Average" if dv=="📈 Average" else "Median"
                pivot = make_pivot(df_week, "GateOut WeekNo", tat_cols, metric)
                st.dataframe(pivot, use_container_width=True, hide_index=True)
            else:
                sub = df_week[df_week["TAT Stage"]==sel_stage].copy()
                render_view_table(sub, dv, extra_cols=["GateOut WeekNo"])
                render_cards(sub, "GateOut WeekNo", sel_stage)

    # ── TAB 4: MONTH WISE ────────────────────────────────────
    with tab_month:
        df_month = time_data.get("month_df", pd.DataFrame())
        if df_month.empty:
            st.info("No GateOut month data available.")
        else:
            st.markdown("**Grouped by Month Name (January, February …)**")
            available = df_month["TAT Stage"].unique().tolist()
            sel_stage = st.selectbox("TAT Stage", ["All Stages"]+available, key="td_month_stage")
            dv = st.radio("View", ["📈 Average","📉 Median","📊 Both"],
                          horizontal=True, key="td_month_view")
            st.markdown("")
            if sel_stage == "All Stages":
                st.markdown("##### 📊 Pivot — Month × TAT Stages")
                metric = "Average" if dv=="📈 Average" else "Median"
                pivot = make_pivot(df_month, "GateOut Month", tat_cols, metric)
                st.dataframe(pivot, use_container_width=True, hide_index=True)
            else:
                sub = df_month[df_month["TAT Stage"]==sel_stage].copy()
                render_view_table(sub, dv, extra_cols=["GateOut Month"])
                render_cards(sub, "GateOut Month", sel_stage)

    # ── ROW COUNT ────────────────────────────────────────────
    st.markdown("---")
    rc1,rc2,rc3 = st.columns(3)
    rc1.metric("Uploaded rows", total_rows)
    rc2.metric("Output rows",   len(result))
    rc3.metric("Match?", "✅ YES" if total_rows==len(result) else "❌ MISMATCH")

    # ── FULL PREVIEW ─────────────────────────────────────────
    st.markdown("---")
    st.subheader(f"👁 Full Data — {len(result)} Rows")
    all_preview = ["Trip ID","Vehicle Number","Transporter Name","Shift",
                   "GateOut Date","GateOut DayOfWeek","GateOut WeekNo","GateOut Month",
                   "YardIn","GateIn","GrossWeight","TareWeight","GateOut",
                   "YI-GI","GI-GW","GW-TW","TW-GO","GI-GO",
                   "GW-LI","LI-LO","LO-TW","GW-GO"]
    prev = [c for c in all_preview if c in result.columns]
    st.dataframe(result[prev], use_container_width=True, height=360)

    # ── DOWNLOAD ─────────────────────────────────────────────
    st.markdown("---")
    buf = export_excel(result, stats_df, groupby_df, time_data, dt_cols, tat_set)
    fname = "Inbound_TAT_Result.xlsx" if mode=="inbound" else "Outbound_TAT_Result.xlsx"
    st.info("✅ Excel has **7 sheets:** Full Data | Overall Stats | Category Stats | Date Wise | Day of Week | Week Wise | Month Wise")
    st.download_button(
        f"⬇️ Download Excel ({len(result)} rows)",
        data=buf, file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, type="primary"
    )


# ─────────────────────────────────────────────────────────────
# SHARED SAVE & RERUN
# ─────────────────────────────────────────────────────────────
def save_and_rerun(result, tat_cols_set, dt_col_names, sel_cat, total_rows, gateout_col):
    tat_cols = list(tat_cols_set)

    # ── ADD ALL TIME DIMENSION COLUMNS ───────────────────────
    if gateout_col and gateout_col != "-- Not Available --":
        dt_go = to_dt(result[gateout_col])
        if dt_go is not None:
            result["GateOut Date"]      = dt_go.dt.strftime("%d-%m-%Y")
            result["GateOut DayOfWeek"] = dt_go.dt.strftime("%A")           # Monday
            result["GateOut WeekNo"]    = dt_go.dt.isocalendar().week.astype(str).apply(
                                            lambda w: f"Week {int(w):02d}")  # Week 01
            result["GateOut Month"]     = dt_go.dt.strftime("%B")           # January
            # Clean NaT
            for c in ["GateOut Date","GateOut DayOfWeek","GateOut WeekNo","GateOut Month"]:
                result[c] = result[c].replace("NaT","")

    # ── BUILD TIME DIMENSION STATS ────────────────────────────
    time_data = {}
    if "GateOut Date" in result.columns:
        time_data["date_df"] = build_time_dimension(
            result, tat_cols, "GateOut Date")
    if "GateOut DayOfWeek" in result.columns:
        time_data["dayofwk_df"] = build_time_dimension(
            result, tat_cols, "GateOut DayOfWeek", sort_order=WEEKDAY_ORDER)
    if "GateOut WeekNo" in result.columns:
        all_weeks = sorted(result["GateOut WeekNo"].unique().tolist())
        time_data["week_df"] = build_time_dimension(
            result, tat_cols, "GateOut WeekNo", sort_order=all_weeks)
    if "GateOut Month" in result.columns:
        time_data["month_df"] = build_time_dimension(
            result, tat_cols, "GateOut Month", sort_order=MONTH_ORDER)

    st.session_state["tat_result"]      = result
    st.session_state["tat_stats"]       = build_stats_table(result, tat_cols)
    st.session_state["tat_groupby"]     = (build_groupby_stats(result, tat_cols, sel_cat)
                                            if sel_cat != "-- None --" else pd.DataFrame())
    st.session_state["tat_time_data"]   = time_data
    st.session_state["tat_dt_cols"]     = dt_col_names
    st.session_state["tat_tat_set"]     = tat_cols_set
    st.session_state["tat_total"]       = total_rows
    st.session_state["tat_gateout_col"] = gateout_col
    st.rerun()


# ═════════════════════════════════════════════════════════════
# INBOUND
# ═════════════════════════════════════════════════════════════
if mode == "inbound":
    st.subheader("📥 Inbound TAT")
    st.markdown("**Process:** `YardIn → GateIn → GrossWeight → TareWeight → GateOut`")
    with st.expander("📖 Stage Reference"):
        st.markdown("""
        | TAT Column | Formula | Meaning |
        |---|---|---|
        | YI-GI | GateIn − YardIn | Yard to Gate Entry |
        | GI-GW | GrossWeight − GateIn | Gate In to Gross Weighment |
        | GW-TW | TareWeight − GrossWeight | Gross to Tare Weighment |
        | TW-GO | GateOut − TareWeight | Tare to Gate Out |
        | GI-GO | GateOut − GateIn | Total Plant Time |
        """)
    st.markdown("---")

    uploaded = st.file_uploader("📂 Upload Inbound Excel", type=["xlsx","xls"], key="ib_upload")
    if uploaded is None:
        st.info("👆 Upload your inbound Excel file to begin.")
        st.session_state.pop("tat_result", None)
        st.stop()

    df = load_file(uploaded)
    total_rows = len(df)
    st.success(f"✅ Loaded: **{total_rows} rows, {len(df.columns)} columns**")
    with st.expander("🔍 Debug Info"):
        st.write("**Columns:**", df.columns.tolist())
        st.dataframe(df.head(3), use_container_width=True)

    st.markdown("---")
    st.subheader("🔧 Map Your Columns")
    all_cols = ["-- Not Available --"] + df.columns.tolist()
    c1, c2 = st.columns(2)
    with c1:
        col_yardin  = st.selectbox("YardIn",      all_cols, index=auto_index(all_cols,"YardIn"),      key="ib_yi")
        col_gatein  = st.selectbox("GateIn",      all_cols, index=auto_index(all_cols,"GateIn"),      key="ib_gi")
        col_grosswt = st.selectbox("GrossWeight", all_cols, index=auto_index(all_cols,"GrossWeight"), key="ib_gw")
    with c2:
        col_tarewt  = st.selectbox("TareWeight",  all_cols, index=auto_index(all_cols,"TareWeight"),  key="ib_tw")
        col_gateout = st.selectbox("GateOut",     all_cols, index=auto_index(all_cols,"GateOut"),     key="ib_go")

    st.markdown("---")
    st.markdown("**🔎 Analysis Options:**")
    cat_possible = ["Transporter Name","Shift","Mat. Group","Material Group",
                    "Unloader Alias","Vehicle Number","Gate Entry Type","Supplier Name"]
    cat_found = ["-- None --"] + [c for c in cat_possible if c in df.columns]
    sel_cat = st.selectbox("📂 Group-by Category (optional)", cat_found, key="ib_cat")
    st.session_state["tat_sel_cat"] = sel_cat
    st.info("📅 Auto-added columns: **GateOut Date · GateOut DayOfWeek · GateOut WeekNo · GateOut Month**")
    st.markdown("---")

    if st.button("⚙️ Calculate Inbound TAT", type="primary", use_container_width=True, key="ib_calc"):
        result = df.copy()
        dt_yardin  = to_dt(result[col_yardin])  if col_yardin  != "-- Not Available --" else None
        dt_gatein  = to_dt(result[col_gatein])  if col_gatein  != "-- Not Available --" else None
        dt_grosswt = to_dt(result[col_grosswt]) if col_grosswt != "-- Not Available --" else None
        dt_tarewt  = to_dt(result[col_tarewt])  if col_tarewt  != "-- Not Available --" else None
        dt_gateout = to_dt(result[col_gateout]) if col_gateout != "-- Not Available --" else None

        st.markdown("#### 📅 DateTime Parse")
        parse_summary([
            ("YardIn",dt_yardin,col_yardin),("GateIn",dt_gatein,col_gatein),
            ("GrossWeight",dt_grosswt,col_grosswt),("TareWeight",dt_tarewt,col_tarewt),
            ("GateOut",dt_gateout,col_gateout),
        ], st.columns(5))

        st.markdown("#### ⚙️ Calculating Stages")
        stages = [
            ("YI-GI",dt_yardin, dt_gatein, "YardIn",     "GateIn"),
            ("GI-GW",dt_gatein, dt_grosswt,"GateIn",     "GrossWeight"),
            ("GW-TW",dt_grosswt,dt_tarewt, "GrossWeight","TareWeight"),
            ("TW-GO",dt_tarewt, dt_gateout,"TareWeight", "GateOut"),
            ("GI-GO",dt_gatein, dt_gateout,"GateIn",     "GateOut"),
        ]
        result, tat_cols_set = calculate_stages(result, stages, st)
        dt_col_names = [c for c in [col_yardin,col_gatein,col_grosswt,col_tarewt,col_gateout]
                        if c != "-- Not Available --"]
        save_and_rerun(result, tat_cols_set, dt_col_names, sel_cat, total_rows, col_gateout)


# ═════════════════════════════════════════════════════════════
# OUTBOUND
# ═════════════════════════════════════════════════════════════
else:
    st.subheader("📤 Outbound TAT")
    st.markdown("**Process:** `YardIn* → YardOut* → GateIn → GrossWeight → LoadingIn → LoadingOut → TareWeight → GateOut`")
    st.markdown("*(ParkIn/ParkOut used as fallback per row when YardIn/YardOut is blank)*")
    with st.expander("📖 Stage Reference"):
        st.markdown("""
        **Full Process:** `YardIn → YardOut → GateIn → GrossWeight → LoadingIn → LoadingOut → TareWeight → GateOut`

        > 💡 **ParkIn Fallback:** If YardIn is blank for a row, ParkIn value is used automatically for that row.
        > Similarly if YardOut is blank, ParkOut is used.

        | TAT Column | Formula | Meaning |
        |---|---|---|
        | YI-GI | GateIn − YardIn* | Yard/Park In to Gate In |
        | YI-YO | YardOut* − YardIn* | Yard In to Yard Out |
        | YO-GI | GateIn − YardOut* | Yard Out to Gate In |
        | GI-GW | GrossWeight − GateIn | Gate In to Gross Weighment |
        | GW-LI | LoadingIn − GrossWeight | Gross Weighment to Loading Start |
        | LI-LO | LoadingOut − LoadingIn | Loading Duration |
        | LO-TW | TareWeight − LoadingOut | Loading End to Tare Weighment |
        | TW-GO | GateOut − TareWeight | Tare to Gate Out |
        | GI-GO | GateOut − GateIn | Total Plant Time |
        | GW-GO | GateOut − GrossWeight | Gross to Gate Out |

        *(ParkIn used as YardIn fallback per row. ParkOut used as YardOut fallback per row.)*
        """)
    st.markdown("---")

    uploaded = st.file_uploader("📂 Upload Outbound Excel", type=["xlsx","xls"], key="ob_upload")
    if uploaded is None:
        st.info("👆 Upload your outbound Excel file to begin.")
        st.session_state.pop("tat_result", None)
        st.stop()

    df = load_file(uploaded)
    total_rows = len(df)
    st.success(f"✅ Loaded: **{total_rows} rows, {len(df.columns)} columns**")
    with st.expander("🔍 Debug Info"):
        st.write("**Columns:**", df.columns.tolist())
        st.dataframe(df.head(3), use_container_width=True)

    st.markdown("---")
    st.subheader("🔧 Map Your Columns")
    all_cols = ["-- Not Available --"] + df.columns.tolist()

    st.markdown("**🕐 Timestamps — map in process order:**")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        col_yardin     = st.selectbox("YardIn",       all_cols, index=auto_index(all_cols,"YardIn"),      key="ob_yi")
        col_parkin     = st.selectbox("ParkIn (fallback for YardIn)", all_cols, index=auto_index(all_cols,"ParkIn"), key="ob_pi")
    with c2:
        col_yardout    = st.selectbox("YardOut",      all_cols, index=auto_index(all_cols,"YardOut"),     key="ob_yo")
        col_parkout    = st.selectbox("ParkOut (fallback for YardOut)", all_cols, index=auto_index(all_cols,"ParkOut"), key="ob_po")
    with c3:
        col_gatein     = st.selectbox("GateIn",       all_cols, index=auto_index(all_cols,"GateIn"),      key="ob_gi")
        col_grosswt    = st.selectbox("GrossWeight",  all_cols, index=auto_index(all_cols,"GrossWeight"), key="ob_gw")
        col_loadingin  = st.selectbox("LoadingIn",    all_cols, index=auto_index(all_cols,"LoadingIn"),   key="ob_li")
    with c4:
        col_loadingout = st.selectbox("LoadingOut",   all_cols, index=auto_index(all_cols,"LoadingOut"),  key="ob_lo")
        col_tarewt     = st.selectbox("TareWeight",   all_cols, index=auto_index(all_cols,"TareWeight"),  key="ob_tw")
        col_gateout    = st.selectbox("GateOut",      all_cols, index=auto_index(all_cols,"GateOut"),     key="ob_go")

    st.info("💡 **ParkIn/ParkOut** are used as fallback per row — only when YardIn/YardOut is blank for that specific row.")

    st.markdown("---")
    st.markdown("**🔎 Analysis Options:**")
    cat_possible = ["Transporter Name","Shift","Mat. Group","Material Group",
                    "Unloader Alias","Vehicle Number","Gate Entry Type","Supplier Name"]
    cat_found = ["-- None --"] + [c for c in cat_possible if c in df.columns]
    sel_cat = st.selectbox("📂 Group-by Category (optional)", cat_found, key="ob_cat")
    st.session_state["tat_sel_cat"] = sel_cat
    st.info("📅 Auto-added columns: **GateOut Date · GateOut DayOfWeek · GateOut WeekNo · GateOut Month**")
    st.markdown("---")

    if st.button("⚙️ Calculate Outbound TAT", type="primary", use_container_width=True, key="ob_calc"):
        result = df.copy()

        # ── Parse all datetime columns ────────────────────────
        dt_yardin     = to_dt(result[col_yardin])     if col_yardin     != "-- Not Available --" else None
        dt_parkin     = to_dt(result[col_parkin])     if col_parkin     != "-- Not Available --" else None
        dt_yardout    = to_dt(result[col_yardout])    if col_yardout    != "-- Not Available --" else None
        dt_parkout    = to_dt(result[col_parkout])    if col_parkout    != "-- Not Available --" else None
        dt_gatein     = to_dt(result[col_gatein])     if col_gatein     != "-- Not Available --" else None
        dt_grosswt    = to_dt(result[col_grosswt])    if col_grosswt    != "-- Not Available --" else None
        dt_loadingin  = to_dt(result[col_loadingin])  if col_loadingin  != "-- Not Available --" else None
        dt_loadingout = to_dt(result[col_loadingout]) if col_loadingout != "-- Not Available --" else None
        dt_tarewt     = to_dt(result[col_tarewt])     if col_tarewt     != "-- Not Available --" else None
        dt_gateout    = to_dt(result[col_gateout])    if col_gateout    != "-- Not Available --" else None

        # ── Apply ParkIn/ParkOut fallback ROW BY ROW ──────────
        # If YardIn is blank for a row → use ParkIn for that row
        if dt_yardin is not None and dt_parkin is not None:
            dt_yardin_eff = dt_yardin.fillna(dt_parkin)
            filled_from_park = int(dt_yardin.isna().sum() - dt_yardin_eff.isna().sum())
            if filled_from_park > 0:
                st.info(f"💡 YardIn: **{filled_from_park} blank rows** filled using ParkIn fallback")
        elif dt_yardin is None and dt_parkin is not None:
            dt_yardin_eff = dt_parkin
            st.info("💡 YardIn not mapped — using ParkIn as YardIn for all rows")
        else:
            dt_yardin_eff = dt_yardin

        # If YardOut is blank for a row → use ParkOut for that row
        if dt_yardout is not None and dt_parkout is not None:
            dt_yardout_eff = dt_yardout.fillna(dt_parkout)
            filled_from_parkout = int(dt_yardout.isna().sum() - dt_yardout_eff.isna().sum())
            if filled_from_parkout > 0:
                st.info(f"💡 YardOut: **{filled_from_parkout} blank rows** filled using ParkOut fallback")
        elif dt_yardout is None and dt_parkout is not None:
            dt_yardout_eff = dt_parkout
            st.info("💡 YardOut not mapped — using ParkOut as YardOut for all rows")
        else:
            dt_yardout_eff = dt_yardout

        # ── Store effective values back for Excel output ───────
        if dt_yardin_eff is not None:
            result["YardIn (Eff)"] = dt_yardin_eff
        if dt_yardout_eff is not None:
            result["YardOut (Eff)"] = dt_yardout_eff

        st.markdown("#### 📅 DateTime Parse Summary")
        parse_summary([
            ("YardIn*",  dt_yardin_eff,  f"{col_yardin}+ParkIn fallback"),
            ("YardOut*", dt_yardout_eff, f"{col_yardout}+ParkOut fallback"),
            ("GateIn",   dt_gatein,      col_gatein),
            ("GrossWt",  dt_grosswt,     col_grosswt),
            ("GateOut",  dt_gateout,     col_gateout),
        ], st.columns(5))

        st.markdown("#### ⚙️ Calculating Stages")
        stages = []

        # YardIn* → YardOut* (YI-YO)
        if dt_yardin_eff is not None and dt_yardout_eff is not None:
            stages.append(("YI-YO", dt_yardin_eff, dt_yardout_eff, "YardIn*", "YardOut*"))

        # YardIn* → GateIn (YI-GI)
        if dt_yardin_eff is not None and dt_gatein is not None:
            stages.append(("YI-GI", dt_yardin_eff, dt_gatein, "YardIn*", "GateIn"))

        # YardOut* → GateIn (YO-GI)
        if dt_yardout_eff is not None and dt_gatein is not None:
            stages.append(("YO-GI", dt_yardout_eff, dt_gatein, "YardOut*", "GateIn"))

        # GateIn → GrossWeight (GI-GW)
        if dt_gatein is not None and dt_grosswt is not None:
            stages.append(("GI-GW", dt_gatein, dt_grosswt, "GateIn", "GrossWeight"))

        # GrossWeight → LoadingIn (GW-LI)
        if dt_grosswt is not None and dt_loadingin is not None:
            stages.append(("GW-LI", dt_grosswt, dt_loadingin, "GrossWeight", "LoadingIn"))

        # LoadingIn → LoadingOut (LI-LO)
        if dt_loadingin is not None and dt_loadingout is not None:
            stages.append(("LI-LO", dt_loadingin, dt_loadingout, "LoadingIn", "LoadingOut"))

        # LoadingOut → TareWeight (LO-TW)
        if dt_loadingout is not None and dt_tarewt is not None:
            stages.append(("LO-TW", dt_loadingout, dt_tarewt, "LoadingOut", "TareWeight"))

        # TareWeight → GateOut (TW-GO)
        if dt_tarewt is not None and dt_gateout is not None:
            stages.append(("TW-GO", dt_tarewt, dt_gateout, "TareWeight", "GateOut"))

        # GateIn → GateOut (GI-GO)
        if dt_gatein is not None and dt_gateout is not None:
            stages.append(("GI-GO", dt_gatein, dt_gateout, "GateIn", "GateOut"))

        # GrossWeight → GateOut (GW-GO)
        if dt_grosswt is not None and dt_gateout is not None:
            stages.append(("GW-GO", dt_grosswt, dt_gateout, "GrossWeight", "GateOut"))

        result, tat_cols_set = calculate_stages(result, stages, st)

        dt_col_names = [c for c in [col_yardin, col_parkin, col_yardout, col_parkout,
                                     col_gatein, col_grosswt, col_loadingin,
                                     col_loadingout, col_tarewt, col_gateout]
                        if c != "-- Not Available --"]
        save_and_rerun(result, tat_cols_set, dt_col_names, sel_cat, total_rows, col_gateout)


# ─────────────────────────────────────────────────────────────
# RENDER (from session_state — survives all button clicks)
# ─────────────────────────────────────────────────────────────
if "tat_result" in st.session_state:
    render_analysis()