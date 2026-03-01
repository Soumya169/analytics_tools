import streamlit as st
import pandas as pd
import sys, os

sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from utils import load_file, hms_to_min, min_to_hms
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Category Analysis", page_icon="📊", layout="wide")

st.title("📊 Category-wise TAT Analysis")
st.markdown("Upload a processed TAT file and analyse TAT by any category — transporter, shift, material, vehicle etc.")
st.markdown("---")

uploaded = st.file_uploader("📂 Upload Processed TAT Excel", type=["xlsx", "xls"], key="cat_upload")
if uploaded is None:
    st.info("👆 Upload output file from **TAT Analysis** (Inbound or Outbound) to begin.")
    st.markdown("""
    **This module expects a file already containing TAT columns like:**
    `YI-GI`, `GI-GW`, `GW-TW`, `TW-GO`, `GI-GO`, `GW-LI`, `LI-LO`, `LO-TW`, `GW-GO`

    Run **TAT Analysis** first → download result → upload here.
    """)
    st.stop()

df = load_file(uploaded)
st.success(f"✅ Loaded: **{len(df)} rows, {len(df.columns)} columns**")

# Detect TAT columns
tat_possible = ["YI-GI", "GI-GW", "GW-TW", "TW-GO", "GI-GO", "GW-LI", "LI-LO", "LO-TW", "TW-GO", "GW-GO"]
tat_found = [c for c in tat_possible if c in df.columns]

cat_possible = ["Transporter Name", "Transporter Code", "Shift", "Mat. Group", "Material Group",
                "Unloader Alias", "Vehicle Number", "Gate Entry Type", "Supplier Name",
                "WT Type", "Ref Doc Type", "Lease/Stock Holder Code"]
cat_found = [c for c in cat_possible if c in df.columns]

if not tat_found:
    st.error("❌ No TAT columns found. Please upload the output file from TAT Analysis module.")
    st.stop()

st.markdown("---")
st.subheader("🔧 Select Analysis Parameters")

a1, a2 = st.columns(2)
with a1:
    sel_tat = st.multiselect("Select TAT Columns to Analyse", tat_found, default=tat_found)
with a2:
    if cat_found:
        sel_cat = st.selectbox("Group By (Category)", cat_found)
    else:
        sel_cat = st.selectbox("Group By (Category)", df.columns.tolist())

st.markdown("---")

if st.button("📊 Run Category Analysis", type="primary", use_container_width=True):

    if not sel_tat:
        st.warning("⚠️ Please select at least one TAT column.")
        st.stop()

    analysis_df = df.copy()

    # Convert HH:MM:SS → minutes
    for col in sel_tat:
        analysis_df[col + "_min"] = analysis_df[col].apply(hms_to_min)

    min_cols = [c + "_min" for c in sel_tat]

    # ── CATEGORY SUMMARY ──────────────────────────────────────
    st.subheader(f"📋 Average TAT by {sel_cat}")

    grp = analysis_df.groupby(sel_cat)[min_cols].mean().round(2).reset_index()
    counts = analysis_df.groupby(sel_cat).size().reset_index(name="Trip Count")
    grp = grp.merge(counts, on=sel_cat)

    # Convert back to HH:MM:SS
    grp_display = grp[[sel_cat, "Trip Count"]].copy()
    for col in sel_tat:
        grp_display[col] = grp[col + "_min"].apply(min_to_hms)

    st.dataframe(grp_display, use_container_width=True)

    # ── OVERALL SUMMARY ───────────────────────────────────────
    st.markdown("---")
    st.subheader("📊 Overall TAT Summary (All Rows)")
    summary_rows = []
    for col in sel_tat:
        vals = analysis_df[col + "_min"].dropna()
        if len(vals) > 0:
            summary_rows.append({
                "TAT Stage": col,
                "Avg": min_to_hms(vals.mean()),
                "Min": min_to_hms(vals.min()),
                "Max": min_to_hms(vals.max()),
                "Median": min_to_hms(vals.median()),
                "Valid Rows": int(vals.count()),
                "Blank Rows": int(len(analysis_df) - vals.count()),
            })
    summary_df = pd.DataFrame(summary_rows)
    st.dataframe(summary_df, use_container_width=True)

    # ── MULTI-CATEGORY BREAKDOWN ──────────────────────────────
    if len(cat_found) > 1:
        st.markdown("---")
        st.subheader("🔀 Cross-Category Breakdown")
        extra_cats = [c for c in cat_found if c != sel_cat]
        sel_extra = st.selectbox("Add second grouping", ["None"] + extra_cats, key="extra_cat")

        if sel_extra != "None":
            cross_grp = analysis_df.groupby([sel_cat, sel_extra])[min_cols].mean().round(2).reset_index()
            cross_cnt = analysis_df.groupby([sel_cat, sel_extra]).size().reset_index(name="Trip Count")
            cross_grp = cross_grp.merge(cross_cnt, on=[sel_cat, sel_extra])
            cross_display = cross_grp[[sel_cat, sel_extra, "Trip Count"]].copy()
            for col in sel_tat:
                cross_display[col] = cross_grp[col + "_min"].apply(min_to_hms)
            st.dataframe(cross_display, use_container_width=True)

    # ── DOWNLOAD ──────────────────────────────────────────────
    st.markdown("---")
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        grp_display.to_excel(writer, index=False, sheet_name="Category Analysis")
        summary_df.to_excel(writer, index=False, sheet_name="Overall Summary")

        for sname in writer.sheets:
            ws = writer.sheets[sname]
            for cell in ws[1]:
                cell.fill = PatternFill("solid", start_color="243F60")
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            ws.freeze_panes = "A2"

    buf.seek(0)
    st.download_button(
        "⬇️ Download Category Analysis Excel",
        data=buf,
        file_name="Category_TAT_Analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, type="primary"
    )