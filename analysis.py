import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="TAT Calculator", page_icon="🚛", layout="wide")

# ─────────────────────────────────────────────────────────────
# SHARED HELPERS
# ─────────────────────────────────────────────────────────────
def to_dt(series):
    try:
        s = series.replace("", pd.NaT)
        return pd.to_datetime(s, dayfirst=True, errors='coerce')
    except:
        return None

def sec_to_hms(sec):
    if pd.isna(sec) or sec < 0:
        return ""
    sec = int(sec)
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def diff_hms(a, b):
    try:
        diff = (b - a).dt.total_seconds()
        return diff.apply(sec_to_hms)
    except:
        return None

def load_file(uploaded):
    file_bytes = uploaded.read()
    df = pd.read_excel(io.BytesIO(file_bytes), keep_default_na=False, na_filter=False)
    df.columns = df.columns.str.strip()
    return df

def build_excel(result, dt_col_names, tat_cols_set):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        result.to_excel(writer, index=False, sheet_name="Result")
        wb = writer.book
        ws = writer.sheets["Result"]

        col_letter_map = {col: get_column_letter(idx) for idx, col in enumerate(result.columns, 1)}

        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        center = Alignment(horizontal="center", vertical="center")
        hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

        # Header row
        for idx, col_name in enumerate(result.columns, 1):
            cell = ws.cell(row=1, column=idx)
            cell.font      = hdr_font
            cell.fill      = PatternFill("solid", start_color="375623") if col_name in tat_cols_set else PatternFill("solid", start_color="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin
        ws.row_dimensions[1].height = 30

        # Data rows
        for row_num in range(2, len(result) + 2):
            bg = PatternFill("solid", start_color="EBF3FB") if row_num % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            for col_idx, col_name in enumerate(result.columns, 1):
                cell = ws.cell(row=row_num, column=col_idx)
                is_tat = col_name in tat_cols_set
                cell.fill      = PatternFill("solid", start_color="E2EFDA") if is_tat else bg
                cell.font      = Font(name="Arial", size=9, bold=is_tat, color="375623" if is_tat else "000000")
                cell.alignment = center
                cell.border    = thin
                if col_name in tat_cols_set:
                    cell.number_format = "@"

        # Datetime columns — store as real Excel datetime
        date_fmt = "DD-MM-YYYY HH:MM:SS"
        for col_name in dt_col_names:
            if col_name in col_letter_map:
                cl = col_letter_map[col_name]
                for row_num in range(2, len(result) + 2):
                    cell = ws[f"{cl}{row_num}"]
                    if isinstance(cell.value, str) and cell.value != "":
                        try:
                            cell.value = pd.to_datetime(cell.value, dayfirst=True).to_pydatetime()
                        except:
                            pass
                    if cell.value and not isinstance(cell.value, str):
                        cell.number_format = date_fmt

        # Column widths
        for idx, col_name in enumerate(result.columns, 1):
            cl = get_column_letter(idx)
            if col_name in tat_cols_set:
                ws.column_dimensions[cl].width = 14
            elif col_name in dt_col_names:
                ws.column_dimensions[cl].width = 22
            else:
                ws.column_dimensions[cl].width = 18

        ws.freeze_panes = "A2"

    buf.seek(0)
    return buf

def auto_index(all_cols, name):
    if name in all_cols:
        return all_cols.index(name)
    lower_map = {c.lower().strip(): i for i, c in enumerate(all_cols)}
    return lower_map.get(name.lower(), 0)

# ─────────────────────────────────────────────────────────────
# SIDEBAR NAVIGATION
# ─────────────────────────────────────────────────────────────
st.sidebar.image("https://img.icons8.com/fluency/96/truck.png", width=60)
st.sidebar.title("🚛 TAT Calculator")
st.sidebar.markdown("---")
page = st.sidebar.radio(
    "📂 Select Module",
    ["🏠 Home", "📥 Inbound TAT", "📤 Outbound TAT", "📊 Category Analysis"],
    index=0
)
st.sidebar.markdown("---")
st.sidebar.markdown("""
**TAT Colour Guide**
🟢 Green cells = Calculated TAT  
🔵 Blue header = Input columns  
🟢 Green header = TAT columns  
""")
st.sidebar.markdown("---")
st.sidebar.markdown("**All TAT values in `HH:MM:SS` format**")

# ═════════════════════════════════════════════════════════════
# PAGE 1 — HOME
# ═════════════════════════════════════════════════════════════
if page == "🏠 Home":
    st.title("🚛 Transport TAT Calculator")
    st.markdown("### Integrated TAT Analysis Platform")
    st.markdown("---")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("""
        <div style='background:#1F4E79;padding:24px;border-radius:12px;text-align:center'>
            <div style='font-size:40px'>📥</div>
            <h3 style='color:white;margin:8px 0'>Inbound TAT</h3>
            <p style='color:#BDD7EE;font-size:13px'>
            Calculates TAT for inbound trips.<br><br>
            <b>Stages:</b><br>
            YardIn → GateIn (YI-GI)<br>
            GateIn → GrossWeight (GI-GW)<br>
            GrossWeight → TareWeight (GW-TW)<br>
            TareWeight → GateOut (TW-GO)<br>
            GateIn → GateOut (GI-GO)
            </p>
        </div>
        """, unsafe_allow_html=True)

    with c2:
        st.markdown("""
        <div style='background:#375623;padding:24px;border-radius:12px;text-align:center'>
            <div style='font-size:40px'>📤</div>
            <h3 style='color:white;margin:8px 0'>Outbound TAT</h3>
            <p style='color:#E2EFDA;font-size:13px'>
            Calculates TAT for outbound trips.<br><br>
            <b>Stages:</b><br>
            GrossWeight → LoadingIn (GW-LI)<br>
            LoadingIn → LoadingOut (LI-LO)<br>
            LoadingOut → TareWeight (LO-TW)<br>
            TareWeight → GateOut (TW-GO)<br>
            GrossWeight → GateOut (GW-GO)
            </p>
        </div>
        """, unsafe_allow_html=True)

    with c3:
        st.markdown("""
        <div style='background:#7B2D8B;padding:24px;border-radius:12px;text-align:center'>
            <div style='font-size:40px'>📊</div>
            <h3 style='color:white;margin:8px 0'>Category Analysis</h3>
            <p style='color:#E9D5F5;font-size:13px'>
            Upload processed TAT file.<br><br>
            Analyse TAT by:<br>
            • Transporter<br>
            • Shift<br>
            • Material Group<br>
            • Vehicle / Unloader
            </p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 📋 How to Use")
    st.markdown("""
    1. Select a module from the **left sidebar**
    2. Upload your Excel file
    3. Map your columns using the dropdowns
    4. Click **Calculate TAT**
    5. Preview the results in the table
    6. Click **Download** to get the fully formatted Excel file
    """)

# ═════════════════════════════════════════════════════════════
# PAGE 2 — INBOUND TAT
# ═════════════════════════════════════════════════════════════
elif page == "📥 Inbound TAT":
    st.title("📥 Inbound TAT Calculator")
    st.markdown("**Process:** YardIn → GateIn → GrossWeight → TareWeight → GateOut")
    st.markdown("---")

    uploaded = st.file_uploader("📂 Upload Inbound Excel File", type=["xlsx","xls"], key="inbound")
    if uploaded is None:
        st.info("👆 Upload your inbound Excel file to begin.")

        st.markdown("""
        | TAT Column | Formula | Meaning |
        |---|---|---|
        | YI-GI | GateIn − YardIn | Time from Yard to Gate Entry |
        | GI-GW | GrossWeight − GateIn | Time from Gate In to Gross Weighment |
        | GW-TW | TareWeight − GrossWeight | Time from Gross to Tare Weighment |
        | TW-GO | GateOut − TareWeight | Time from Tare Weighment to Gate Out |
        | GI-GO | GateOut − GateIn | Total Plant Processing Time |
        """)
        st.stop()

    df = load_file(uploaded)
    total_rows = len(df)
    st.success(f"✅ Loaded: **{total_rows} rows, {len(df.columns)} columns**")

    with st.expander("🔍 Debug Info"):
        st.write("**Columns:**", df.columns.tolist())
        st.dataframe(df.head(3), use_container_width=True)

    st.markdown("---")
    st.subheader("🔧 Map Columns")
    all_cols = ["-- Not Available --"] + df.columns.tolist()

    c1, c2 = st.columns(2)
    with c1:
        col_yardin  = st.selectbox("YardIn",      all_cols, index=auto_index(all_cols,"YardIn"))
        col_gatein  = st.selectbox("GateIn",      all_cols, index=auto_index(all_cols,"GateIn"))
        col_grosswt = st.selectbox("GrossWeight", all_cols, index=auto_index(all_cols,"GrossWeight"))
    with c2:
        col_tarewt  = st.selectbox("TareWeight",  all_cols, index=auto_index(all_cols,"TareWeight"))
        col_gateout = st.selectbox("GateOut",     all_cols, index=auto_index(all_cols,"GateOut"))

    st.markdown("---")

    if st.button("⚙️ Calculate Inbound TAT", type="primary", use_container_width=True):
        result = df.copy()
        errors = []

        dt_yardin  = to_dt(result[col_yardin])  if col_yardin  != "-- Not Available --" else None
        dt_gatein  = to_dt(result[col_gatein])  if col_gatein  != "-- Not Available --" else None
        dt_grosswt = to_dt(result[col_grosswt]) if col_grosswt != "-- Not Available --" else None
        dt_tarewt  = to_dt(result[col_tarewt])  if col_tarewt  != "-- Not Available --" else None
        dt_gateout = to_dt(result[col_gateout]) if col_gateout != "-- Not Available --" else None

        # Parse summary
        st.markdown("#### 📅 DateTime Parse")
        p1,p2,p3,p4,p5 = st.columns(5)
        for w,lbl,dts,cn in [(p1,"YardIn",dt_yardin,col_yardin),(p2,"GateIn",dt_gatein,col_gatein),
                              (p3,"GrossWt",dt_grosswt,col_grosswt),(p4,"TareWt",dt_tarewt,col_tarewt),
                              (p5,"GateOut",dt_gateout,col_gateout)]:
            w.metric(lbl, f"{int(dts.notna().sum())}/{len(dts)}" if dts is not None else "Not mapped", f"← {cn}" if dts is not None else "")

        st.markdown("#### ⚙️ Results")
        tat_cols_set = set()
        stages = [
            ("YI-GI", dt_yardin,  dt_gatein,  "YardIn",      "GateIn"),
            ("GI-GW", dt_gatein,  dt_grosswt, "GateIn",      "GrossWeight"),
            ("GW-TW", dt_grosswt, dt_tarewt,  "GrossWeight", "TareWeight"),
            ("TW-GO", dt_tarewt,  dt_gateout, "TareWeight",  "GateOut"),
            ("GI-GO", dt_gatein,  dt_gateout, "GateIn",      "GateOut"),
        ]
        for col_name, dt_a, dt_b, from_l, to_l in stages:
            if dt_a is not None and dt_b is not None:
                val = diff_hms(dt_a, dt_b)
                if val is not None:
                    result[col_name] = val
                    tat_cols_set.add(col_name)
                    filled = int((val != "").sum())
                    sample = val[val != ""].iloc[0] if filled > 0 else "N/A"
                    st.success(f"✅ **{col_name}** ({to_l} − {from_l}) → {filled}/{len(result)} rows | Sample: `{sample}`")
            else:
                errors.append(f"{col_name}: {from_l} or {to_l} not mapped")

        if errors:
            for e in errors: st.warning(f"⚠️ {e}")

        # Row check
        rc1,rc2,rc3 = st.columns(3)
        rc1.metric("Uploaded", total_rows)
        rc2.metric("Output",   len(result))
        rc3.metric("Match?", "✅ YES" if total_rows==len(result) else "❌ NO")

        # Preview
        st.subheader(f"👁 Preview — {len(result)} Rows")
        prev = [c for c in ["Trip ID","Vehicle Number","Transporter Name",
                             "YardIn","GateIn","GrossWeight","TareWeight","GateOut",
                             "YI-GI","GI-GW","GW-TW","TW-GO","GI-GO"] if c in result.columns]
        st.dataframe(result[prev], use_container_width=True, height=400)

        # Download
        dt_col_names = [c for c in [col_yardin,col_gatein,col_grosswt,col_tarewt,col_gateout] if c != "-- Not Available --"]
        buf = build_excel(result, dt_col_names, tat_cols_set)
        st.info(f"✅ Excel ready — **{len(result)} rows** | Datetimes as `DD-MM-YYYY HH:MM:SS` | TAT as `HH:MM:SS`")
        st.download_button(
            f"⬇️ Download Inbound Excel ({len(result)} rows)",
            data=buf,
            file_name="Inbound_TAT_Result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary"
        )

# ═════════════════════════════════════════════════════════════
# PAGE 3 — OUTBOUND TAT
# ═════════════════════════════════════════════════════════════
elif page == "📤 Outbound TAT":
    st.title("📤 Outbound TAT Calculator")
    st.markdown("**Process:** GrossWeight → LoadingIn → LoadingOut → TareWeight → GateOut")
    st.markdown("---")

    uploaded = st.file_uploader("📂 Upload Outbound Excel File", type=["xlsx","xls"], key="outbound")
    if uploaded is None:
        st.info("👆 Upload your outbound Excel file to begin.")
        st.markdown("""
        | TAT Column | Formula | Meaning |
        |---|---|---|
        | GW-LI | LoadingIn − GrossWeight | Time from Gross Weighment to Loading Start |
        | LI-LO | LoadingOut − LoadingIn | Loading Duration |
        | LO-TW | TareWeight − LoadingOut | Time from Loading End to Tare Weighment |
        | TW-GO | GateOut − TareWeight | Time from Tare to Gate Out |
        | GW-GO | GateOut − GrossWeight | Total Outbound Plant Time |
        """)
        st.stop()

    df = load_file(uploaded)
    total_rows = len(df)
    st.success(f"✅ Loaded: **{total_rows} rows, {len(df.columns)} columns**")

    with st.expander("🔍 Debug Info"):
        st.write("**Columns:**", df.columns.tolist())
        st.dataframe(df.head(3), use_container_width=True)

    st.markdown("---")
    st.subheader("🔧 Map Columns")
    all_cols = ["-- Not Available --"] + df.columns.tolist()

    c1, c2 = st.columns(2)
    with c1:
        col_grosswt   = st.selectbox("GrossWeight (datetime)",  all_cols, index=auto_index(all_cols,"GrossWeight"))
        col_loadingin = st.selectbox("LoadingIn (datetime)",    all_cols, index=auto_index(all_cols,"LoadingIn"))
        col_loadingout= st.selectbox("LoadingOut (datetime)",   all_cols, index=auto_index(all_cols,"LoadingOut"))
    with c2:
        col_tarewt    = st.selectbox("TareWeight (datetime)",   all_cols, index=auto_index(all_cols,"TareWeight"))
        col_gateout   = st.selectbox("GateOut (datetime)",      all_cols, index=auto_index(all_cols,"GateOut"))

    # Optional: YardIn and GateIn for YI-GI and GI-GW if available
    st.markdown("**Optional — if YardIn & GateIn available:**")
    o1, o2 = st.columns(2)
    with o1:
        col_yardin = st.selectbox("YardIn (optional)", all_cols, index=auto_index(all_cols,"YardIn"))
    with o2:
        col_gatein = st.selectbox("GateIn (optional)", all_cols, index=auto_index(all_cols,"GateIn"))

    st.markdown("---")

    if st.button("⚙️ Calculate Outbound TAT", type="primary", use_container_width=True):
        result = df.copy()
        errors = []

        dt_yardin    = to_dt(result[col_yardin])    if col_yardin    != "-- Not Available --" else None
        dt_gatein    = to_dt(result[col_gatein])    if col_gatein    != "-- Not Available --" else None
        dt_grosswt   = to_dt(result[col_grosswt])   if col_grosswt   != "-- Not Available --" else None
        dt_loadingin = to_dt(result[col_loadingin]) if col_loadingin != "-- Not Available --" else None
        dt_loadingout= to_dt(result[col_loadingout])if col_loadingout!= "-- Not Available --" else None
        dt_tarewt    = to_dt(result[col_tarewt])    if col_tarewt    != "-- Not Available --" else None
        dt_gateout   = to_dt(result[col_gateout])   if col_gateout   != "-- Not Available --" else None

        # Parse summary
        st.markdown("#### 📅 DateTime Parse")
        cols_display = [
            ("GrossWt",    dt_grosswt,    col_grosswt),
            ("LoadingIn",  dt_loadingin,  col_loadingin),
            ("LoadingOut", dt_loadingout, col_loadingout),
            ("TareWt",     dt_tarewt,     col_tarewt),
            ("GateOut",    dt_gateout,    col_gateout),
        ]
        p_cols = st.columns(len(cols_display))
        for w, (lbl, dts, cn) in zip(p_cols, cols_display):
            w.metric(lbl, f"{int(dts.notna().sum())}/{len(dts)}" if dts is not None else "Not mapped", f"← {cn}" if dts is not None else "")

        st.markdown("#### ⚙️ Results")
        tat_cols_set = set()

        # Outbound stages
        outbound_stages = [
            ("GW-LI", dt_grosswt,    dt_loadingin,  "GrossWeight", "LoadingIn"),
            ("LI-LO", dt_loadingin,  dt_loadingout, "LoadingIn",   "LoadingOut"),
            ("LO-TW", dt_loadingout, dt_tarewt,     "LoadingOut",  "TareWeight"),
            ("TW-GO", dt_tarewt,     dt_gateout,    "TareWeight",  "GateOut"),
            ("GW-GO", dt_grosswt,    dt_gateout,    "GrossWeight", "GateOut"),
        ]

        # Optional inbound pre-stages
        if dt_yardin is not None and dt_gatein is not None:
            outbound_stages.insert(0, ("YI-GI", dt_yardin, dt_gatein, "YardIn", "GateIn"))
        if dt_gatein is not None and dt_grosswt is not None:
            outbound_stages.insert(1 if dt_yardin is not None else 0,
                                   ("GI-GW", dt_gatein, dt_grosswt, "GateIn", "GrossWeight"))

        for col_name, dt_a, dt_b, from_l, to_l in outbound_stages:
            if dt_a is not None and dt_b is not None:
                val = diff_hms(dt_a, dt_b)
                if val is not None:
                    result[col_name] = val
                    tat_cols_set.add(col_name)
                    filled = int((val != "").sum())
                    sample = val[val != ""].iloc[0] if filled > 0 else "N/A"
                    st.success(f"✅ **{col_name}** ({to_l} − {from_l}) → {filled}/{len(result)} rows | Sample: `{sample}`")
            else:
                errors.append(f"{col_name}: {from_l} or {to_l} not mapped")

        if errors:
            for e in errors: st.warning(f"⚠️ {e}")

        rc1,rc2,rc3 = st.columns(3)
        rc1.metric("Uploaded", total_rows)
        rc2.metric("Output",   len(result))
        rc3.metric("Match?", "✅ YES" if total_rows==len(result) else "❌ NO")

        st.subheader(f"👁 Preview — {len(result)} Rows")
        prev = [c for c in ["Trip ID","Vehicle Number","Transporter Name",
                             "GrossWeight","LoadingIn","LoadingOut","TareWeight","GateOut",
                             "GW-LI","LI-LO","LO-TW","TW-GO","GW-GO",
                             "YI-GI","GI-GW"] if c in result.columns]
        st.dataframe(result[prev], use_container_width=True, height=400)

        dt_col_names = [c for c in [col_yardin,col_gatein,col_grosswt,col_loadingin,
                                     col_loadingout,col_tarewt,col_gateout] if c != "-- Not Available --"]
        buf = build_excel(result, dt_col_names, tat_cols_set)
        st.info(f"✅ Excel ready — **{len(result)} rows** | Datetimes as `DD-MM-YYYY HH:MM:SS` | TAT as `HH:MM:SS`")
        st.download_button(
            f"⬇️ Download Outbound Excel ({len(result)} rows)",
            data=buf,
            file_name="Outbound_TAT_Result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary"
        )

# ═════════════════════════════════════════════════════════════
# PAGE 4 — CATEGORY ANALYSIS
# ═════════════════════════════════════════════════════════════
elif page == "📊 Category Analysis":
    st.title("📊 Category-wise TAT Analysis")
    st.markdown("Upload a processed TAT file (Inbound or Outbound) to analyse TAT by category.")
    st.markdown("---")

    uploaded = st.file_uploader("📂 Upload Processed TAT Excel", type=["xlsx","xls"], key="analysis")
    if uploaded is None:
        st.info("👆 Upload a processed TAT file (output from Inbound or Outbound module).")
        st.stop()

    df = load_file(uploaded)
    st.success(f"✅ Loaded: **{len(df)} rows, {len(df.columns)} columns**")

    # Detect TAT columns (HH:MM:SS text)
    tat_possible = ["YI-GI","GI-GW","GW-TW","TW-GO","GI-GO","GW-LI","LI-LO","LO-TW","GW-GO"]
    tat_found = [c for c in tat_possible if c in df.columns]

    # Category columns
    cat_possible = ["Transporter Name","Shift","Mat. Group","Material Group",
                    "Unloader Alias","Vehicle Number","Gate Entry Type","Supplier Name","WT Type"]
    cat_found = [c for c in cat_possible if c in df.columns]

    if not tat_found:
        st.error("❌ No TAT columns found (YI-GI, GI-GW etc.). Please upload output from Inbound/Outbound module.")
        st.stop()

    st.markdown("---")
    st.subheader("🔧 Select Analysis Parameters")

    a1, a2 = st.columns(2)
    with a1:
        sel_tat = st.multiselect("Select TAT Columns to Analyse", tat_found, default=tat_found)
    with a2:
        sel_cat = st.selectbox("Group By (Category)", cat_found if cat_found else df.columns.tolist())

    st.markdown("---")

    if st.button("📊 Run Analysis", type="primary", use_container_width=True):

        # Convert HH:MM:SS → total minutes for analysis
        def hms_to_min(val):
            try:
                parts = str(val).split(":")
                if len(parts) == 3:
                    return int(parts[0])*60 + int(parts[1]) + int(parts[2])/60
                return None
            except:
                return None

        analysis_df = df.copy()
        for col in sel_tat:
            analysis_df[col + "_min"] = analysis_df[col].apply(hms_to_min)

        min_cols = [c + "_min" for c in sel_tat]

        if sel_cat not in analysis_df.columns:
            st.error(f"Category column '{sel_cat}' not found.")
            st.stop()

        grp = analysis_df.groupby(sel_cat)[min_cols].mean().round(2).reset_index()
        grp.columns = [sel_cat] + sel_tat

        # Convert back to HH:MM:SS for display
        def min_to_hms(m):
            if pd.isna(m): return "–"
            total_sec = int(m * 60)
            h = total_sec // 3600
            mn = (total_sec % 3600) // 60
            s = total_sec % 60
            return f"{h:02d}:{mn:02d}:{s:02d}"

        grp_display = grp.copy()
        for col in sel_tat:
            grp_display[col] = grp_display[col].apply(min_to_hms)

        # Count of trips per category
        counts = analysis_df.groupby(sel_cat).size().reset_index(name="Trip Count")
        grp_display = grp_display.merge(counts, on=sel_cat)

        st.subheader(f"📋 Average TAT by {sel_cat}")
        st.dataframe(grp_display, use_container_width=True)

        # Overall summary
        st.markdown("---")
        st.subheader("📊 Overall TAT Summary")
        summary_rows = []
        for col in sel_tat:
            vals = analysis_df[col + "_min"].dropna()
            if len(vals) > 0:
                summary_rows.append({
                    "TAT Stage": col,
                    "Avg":  min_to_hms(vals.mean()),
                    "Min":  min_to_hms(vals.min()),
                    "Max":  min_to_hms(vals.max()),
                    "Valid Rows": int(vals.count()),
                    "Blank Rows": int(len(analysis_df) - vals.count()),
                })
        st.dataframe(pd.DataFrame(summary_rows), use_container_width=True)

        # Download analysis
        st.markdown("---")
        buf2 = io.BytesIO()
        with pd.ExcelWriter(buf2, engine='openpyxl') as writer:
            grp_display.to_excel(writer, index=False, sheet_name="Category Analysis")
            pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="Overall Summary")

            for sheet_name in ["Category Analysis","Overall Summary"]:
                ws2 = writer.sheets[sheet_name]
                hfill = PatternFill("solid", start_color="7B2D8B")
                hfont = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                for cell in ws2[1]:
                    cell.fill = hfill
                    cell.font = hfont
                    cell.alignment = Alignment(horizontal="center")
                for col_idx in range(1, ws2.max_column + 1):
                    ws2.column_dimensions[get_column_letter(col_idx)].width = 20
                ws2.freeze_panes = "A2"

        buf2.seek(0)
        st.download_button(
            "⬇️ Download Category Analysis Excel",
            data=buf2,
            file_name="Category_TAT_Analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary"
        )