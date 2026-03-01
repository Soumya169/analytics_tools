import streamlit as st
import pandas as pd
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from utils import load_file, to_dt, auto_index, build_excel, hms_to_min, min_to_hms
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Loader Analysis", page_icon="🏗️", layout="wide")

st.title("🏗️ Loader Analysis")
st.markdown("Analyse loader-wise performance — trip count, loading duration, material handled.")
st.markdown("---")

# ── INBOUND / OUTBOUND TOGGLE ────────────────────────────────
col_btn1, col_btn2, _ = st.columns([1, 1, 4])
with col_btn1:
    ib = st.button("📥 Inbound", use_container_width=True,
                   type="primary" if st.session_state.get("loader_mode","inbound") == "inbound" else "secondary")
with col_btn2:
    ob = st.button("📤 Outbound", use_container_width=True,
                   type="primary" if st.session_state.get("loader_mode","inbound") == "outbound" else "secondary")
if ib: st.session_state["loader_mode"] = "inbound"
if ob: st.session_state["loader_mode"] = "outbound"
mode = st.session_state.get("loader_mode", "inbound")

st.markdown(f"**Mode: {'📥 Inbound' if mode == 'inbound' else '📤 Outbound'}**")
st.markdown("---")

uploaded = st.file_uploader("📂 Upload Excel File", type=["xlsx","xls"], key="loader_upload")
if uploaded is None:
    st.info("👆 Upload your Excel file to begin Loader Analysis.")
    st.markdown("""
    **Expected columns for Loader Analysis:**
    - `Loader Name` or `Unloader Alias` — loader identifier
    - `LoadingIn` / `LoadingOut` — loading start & end datetime
    - `Shift` — shift (A/B/C)
    - `Mat. Group` or `Material Group` — material category
    - `Challan Quantity` or `Net Weight` — quantity loaded
    """)
    st.stop()

df = load_file(uploaded)
total_rows = len(df)
st.success(f"✅ Loaded: **{total_rows} rows, {len(df.columns)} columns**")

with st.expander("🔍 Debug Info"):
    st.write("**Columns:**", df.columns.tolist())
    st.dataframe(df.head(3), use_container_width=True)

st.markdown("---")
st.subheader("🔧 Map Your Columns")
all_cols = ["-- Not Available --"] + df.columns.tolist()

c1, c2, c3 = st.columns(3)
with c1:
    col_loader     = st.selectbox("Loader / Unloader Name", all_cols,
                                  index=auto_index(all_cols,"Unloader Alias"))
    col_loadingin  = st.selectbox("Loading In (datetime)",  all_cols,
                                  index=auto_index(all_cols,"LoadingIn"))
with c2:
    col_loadingout = st.selectbox("Loading Out (datetime)", all_cols,
                                  index=auto_index(all_cols,"LoadingOut"))
    col_shift      = st.selectbox("Shift",                  all_cols,
                                  index=auto_index(all_cols,"Shift"))
with c3:
    col_matgrp     = st.selectbox("Material Group",         all_cols,
                                  index=auto_index(all_cols,"Mat. Group"))
    col_qty        = st.selectbox("Quantity / Net Weight",  all_cols,
                                  index=auto_index(all_cols,"Challan Quantity"))

st.markdown("---")

if st.button("📊 Run Loader Analysis", type="primary", use_container_width=True):

    result = df.copy()

    # Calculate Loading Duration (LI-LO) if both columns mapped
    if col_loadingin != "-- Not Available --" and col_loadingout != "-- Not Available --":
        dt_li = to_dt(result[col_loadingin])
        dt_lo = to_dt(result[col_loadingout])
        if dt_li is not None and dt_lo is not None:
            diff = (dt_lo - dt_li).dt.total_seconds()
            from utils import sec_to_hms
            result["LI-LO (Duration)"] = diff.apply(sec_to_hms)
            result["LI-LO_min"] = diff / 60
            st.success(f"✅ Loading Duration (LI-LO) calculated for {int((diff >= 0).sum())} rows")

    st.markdown("---")

    # ── SUMMARY TABLE ─────────────────────────────────────────
    if col_loader != "-- Not Available --":
        st.subheader(f"📋 Loader-wise Summary")

        grp_cols = [col_loader]
        if col_shift  != "-- Not Available --": grp_cols.append(col_shift)
        if col_matgrp != "-- Not Available --": grp_cols.append(col_matgrp)

        agg_dict = {"Trip Count": (col_loader, "count")}
        if "LI-LO_min" in result.columns:
            agg_dict["Avg Loading Duration (min)"] = ("LI-LO_min", "mean")
            agg_dict["Total Loading Time (min)"]   = ("LI-LO_min", "sum")
        if col_qty != "-- Not Available --":
            result[col_qty] = pd.to_numeric(result[col_qty], errors='coerce')
            agg_dict["Total Quantity"] = (col_qty, "sum")
            agg_dict["Avg Quantity"]   = (col_qty, "mean")

        summary = result.groupby(col_loader).agg(**{
            k: v for k, v in agg_dict.items()
        }).round(2).reset_index()

        # Convert avg duration back to HH:MM:SS
        if "Avg Loading Duration (min)" in summary.columns:
            summary["Avg Duration (HH:MM:SS)"] = summary["Avg Loading Duration (min)"].apply(min_to_hms)
            summary = summary.drop(columns=["Avg Loading Duration (min)"])

        st.dataframe(summary, use_container_width=True)

        # Shift-wise breakdown
        if col_shift != "-- Not Available --":
            st.markdown("---")
            st.subheader("🕐 Shift-wise Loader Performance")
            shift_grp = result.groupby([col_loader, col_shift]).size().reset_index(name="Trip Count")
            st.dataframe(shift_grp, use_container_width=True)

        # Material-wise breakdown
        if col_matgrp != "-- Not Available --":
            st.markdown("---")
            st.subheader("📦 Material-wise Loader Performance")
            mat_grp = result.groupby([col_loader, col_matgrp]).size().reset_index(name="Trip Count")
            if col_qty != "-- Not Available --":
                mat_qty = result.groupby([col_loader, col_matgrp])[col_qty].sum().round(2).reset_index()
                mat_grp = mat_grp.merge(mat_qty, on=[col_loader, col_matgrp])
                mat_grp.columns = [col_loader, col_matgrp, "Trip Count", "Total Quantity"]
            st.dataframe(mat_grp, use_container_width=True)

    else:
        st.warning("⚠️ Please map the Loader/Unloader Name column to see summary.")

    # Full data preview
    st.markdown("---")
    st.subheader(f"👁 Full Data — {len(result)} Rows")
    st.dataframe(result, use_container_width=True, height=400)

    # Download
    st.markdown("---")
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        result.to_excel(writer, index=False, sheet_name="Loader Data")
        if col_loader != "-- Not Available --":
            summary.to_excel(writer, index=False, sheet_name="Loader Summary")

        for sname in writer.sheets:
            ws = writer.sheets[sname]
            for cell in ws[1]:
                cell.fill = PatternFill("solid", start_color="375623")
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            ws.freeze_panes = "A2"

    buf.seek(0)
    st.download_button(
        f"⬇️ Download Loader Analysis Excel ({len(result)} rows)",
        data=buf,
        file_name="Loader_Analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, type="primary"
    )