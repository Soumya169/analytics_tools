import streamlit as st
import pandas as pd
import sys, os, json
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from utils import load_file, to_dt, auto_index, calculate_stages, hms_to_min, min_to_hms
import io
import streamlit.components.v1 as components
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="TAT Intelligence", page_icon="⚡", layout="wide")

st.markdown("""
<style>
#MainMenu,footer,header{visibility:hidden}
.block-container{padding:0.5rem 0 0!important;max-width:100%!important}
[data-testid="stSidebar"]{background:#0a0f1e!important;border-right:1px solid rgba(99,102,241,0.15)!important}
[data-testid="stSidebar"] .block-container{padding:1rem!important}
body,[class*="css"],.stApp{background:#060912!important;color:#e2e8f0!important}
[data-testid="stFileUploaderDropzone"]{background:rgba(99,102,241,0.03)!important;border:2px dashed rgba(99,102,241,0.25)!important;border-radius:12px!important}
[data-testid="stSelectbox"]>div>div{background:#0d1117!important;border:1px solid rgba(99,102,241,0.2)!important;border-radius:9px!important;color:#e2e8f0!important}
[data-testid="stSelectbox"] label{color:#64748b!important;font-size:0.72rem!important;text-transform:uppercase;letter-spacing:0.06em}
.stRadio label{color:#94a3b8!important;font-size:0.85rem!important}
.stButton>button{background:linear-gradient(135deg,#6366f1,#8b5cf6)!important;color:#fff!important;border:none!important;border-radius:10px!important;font-weight:600!important;padding:0.55rem 1.5rem!important;box-shadow:0 0 20px rgba(99,102,241,0.35)!important;transition:all 0.3s!important}
.stButton>button:hover{transform:translateY(-2px)!important;box-shadow:0 0 35px rgba(99,102,241,0.55)!important}
.stDownloadButton>button{background:rgba(16,185,129,0.15)!important;color:#34d399!important;border:1px solid rgba(16,185,129,0.3)!important;border-radius:10px!important;font-weight:600!important}
.stDownloadButton>button:hover{background:rgba(16,185,129,0.25)!important}
[data-testid="stAlert"]{background:rgba(99,102,241,0.07)!important;border:1px solid rgba(99,102,241,0.2)!important;border-radius:10px!important}
.stSuccess{background:rgba(16,185,129,0.08)!important;border:1px solid rgba(16,185,129,0.25)!important;color:#34d399!important;border-radius:10px!important}
</style>
""", unsafe_allow_html=True)

WEEKDAY_ORDER = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
MONTH_ORDER   = ["January","February","March","April","May","June","July","August","September","October","November","December"]

# ─────────────────────────────────────────────────────────────
# DATA PROCESSING HELPERS  (logic unchanged)
# ─────────────────────────────────────────────────────────────
def hms_str_to_fraction(v):
    try:
        if not v or v=="–" or pd.isna(v): return None
        p=str(v).strip().split(":")
        if len(p)!=3: return None
        return (int(p[0])*3600+int(p[1])*60+int(p[2]))/86400
    except: return None

def build_stats(result, tat_cols):
    rows=[]
    for col in tat_cols:
        if col not in result.columns: continue
        mins=result[col].apply(hms_to_min).dropna()
        if len(mins)==0: continue
        rows.append({"stage":col,"avg":min_to_hms(mins.mean()),"med":min_to_hms(mins.median()),
                     "min":min_to_hms(mins.min()),"max":min_to_hms(mins.max()),
                     "avg_min":round(float(mins.mean()),2),"med_min":round(float(mins.median()),2),
                     "min_min":round(float(mins.min()),2),"max_min":round(float(mins.max()),2),
                     "valid":int(mins.count()),"blank":int(len(result)-mins.count())})
    return rows

def build_groupby(result, tat_cols, group_col):
    if not group_col or group_col not in result.columns: return []
    rows=[]; tmp=result.copy()
    for col in tat_cols:
        if col in tmp.columns: tmp[col+"_m"]=tmp[col].apply(hms_to_min)
    for col in tat_cols:
        mc=col+"_m"
        if mc not in tmp.columns: continue
        grp=tmp.groupby(group_col)[mc]
        avg,med,cnt=grp.mean(),grp.median(),grp.count()
        for cat in avg.index:
            rows.append({"category":str(cat),"stage":col,
                         "avg":min_to_hms(avg[cat]),"med":min_to_hms(med[cat]),
                         "avg_min":round(float(avg[cat]),2),"med_min":round(float(med[cat]),2),
                         "count":int(cnt[cat])})
    return rows

def build_timedim(result, tat_cols, group_col, sort_order=None):
    if group_col not in result.columns: return []
    rows=[]; tmp=result.copy()
    for col in tat_cols:
        if col in tmp.columns: tmp[col+"_m"]=tmp[col].apply(hms_to_min)
    for col in tat_cols:
        mc=col+"_m"
        if mc not in tmp.columns: continue
        grp=tmp.groupby(group_col)[mc]
        avg=grp.mean();med=grp.median();mn=grp.min();mx=grp.max();cnt=grp.count()
        keys=[k for k in (sort_order or []) if k in avg.index]
        keys+=[k for k in avg.index if k not in keys]
        for k in keys:
            rows.append({"label":str(k),"stage":col,"count":int(cnt[k]),
                         "avg":min_to_hms(avg[k]),"med":min_to_hms(med[k]),
                         "min":min_to_hms(mn[k]),"max":min_to_hms(mx[k]),
                         "avg_min":round(float(avg[k]),2),"med_min":round(float(med[k]),2),
                         "min_min":round(float(mn[k]),2),"max_min":round(float(mx[k]),2)})
    return rows

def add_time_cols(result, gateout_col):
    if not gateout_col or gateout_col=="-- Not Available --": return result
    dt_go=to_dt(result[gateout_col])
    if dt_go is None: return result
    result["GateOut Date"]=dt_go.dt.strftime("%d-%m-%Y")
    result["GateOut DayOfWeek"]=dt_go.dt.strftime("%A")
    result["GateOut WeekNo"]=dt_go.dt.isocalendar().week.astype(str).apply(lambda w:f"Week {int(w):02d}")
    result["GateOut Month"]=dt_go.dt.strftime("%B")
    for c in ["GateOut Date","GateOut DayOfWeek","GateOut WeekNo","GateOut Month"]:
        result[c]=result[c].replace("NaT","")
    return result

def save_state(result, tat_cols_set, dt_col_names, sel_cat, total_rows, gateout_col):
    tat_cols=list(tat_cols_set)
    result=add_time_cols(result, gateout_col)
    td={}
    if "GateOut Date"      in result.columns: td["date"] =build_timedim(result,tat_cols,"GateOut Date")
    if "GateOut DayOfWeek" in result.columns: td["dow"]  =build_timedim(result,tat_cols,"GateOut DayOfWeek",WEEKDAY_ORDER)
    if "GateOut WeekNo"    in result.columns: td["week"] =build_timedim(result,tat_cols,"GateOut WeekNo",sorted(result["GateOut WeekNo"].unique().tolist()))
    if "GateOut Month"     in result.columns: td["month"]=build_timedim(result,tat_cols,"GateOut Month",MONTH_ORDER)
    st.session_state.update({
        "tat_result":result,"tat_stats":build_stats(result,tat_cols),
        "tat_groupby":build_groupby(result,tat_cols,sel_cat) if sel_cat!="-- None --" else [],
        "tat_time":td,"tat_dt_cols":dt_col_names,"tat_tat_set":tat_cols_set,
        "tat_total":total_rows,"tat_gateout_col":gateout_col,"tat_sel_cat":sel_cat
    })
    st.rerun()

# ─────────────────────────────────────────────────────────────
# EXCEL EXPORT  (logic unchanged)
# ─────────────────────────────────────────────────────────────
def style_ws(ws,hc):
    thin=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.fill=PatternFill("solid",start_color=hc);cell.font=Font(name="Arial",size=10,bold=True,color="FFFFFF")
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);cell.border=thin
    ws.row_dimensions[1].height=28
    for rn in range(2,ws.max_row+1):
        bg="EBF3FB" if rn%2==0 else "FFFFFF"
        for cell in ws[rn]:
            cell.alignment=Alignment(horizontal="center",vertical="center")
            cell.fill=PatternFill("solid",start_color=bg);cell.border=thin
    for ci in range(1,ws.max_column+1): ws.column_dimensions[get_column_letter(ci)].width=20
    ws.freeze_panes="A2"

def fmt_time_ws(ws,mr):
    TC={"Average","Median","Min","Max"}
    headers=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
    for ci,h in enumerate(headers,1):
        if h in TC:
            for rn in range(2,mr+1):
                cell=ws.cell(rn,ci);frac=hms_str_to_fraction(cell.value)
                if frac is not None: cell.value=frac;cell.number_format="[HH]:MM:SS"
                cell.alignment=Alignment(horizontal="center",vertical="center")

def export_excel(result,stats_rows,groupby_rows,time_data,dt_col_names,tat_cols_set):
    def s2df(rows):
        if not rows: return pd.DataFrame()
        return pd.DataFrame([{"TAT Stage":r["stage"],"Average":r["avg"],"Median":r["med"],
                               "Min":r["min"],"Max":r["max"],"Valid Rows":r["valid"],"Blank Rows":r["blank"]} for r in rows])
    def g2df(rows):
        if not rows: return pd.DataFrame()
        return pd.DataFrame([{"Category":r["category"],"TAT Stage":r["stage"],"Average":r["avg"],
                               "Median":r["med"],"Trip Count":r["count"]} for r in rows])
    def t2df(rows,gcol):
        if not rows: return pd.DataFrame()
        return pd.DataFrame([{gcol:r["label"],"TAT Stage":r["stage"],"Trip Count":r["count"],
                               "Average":r["avg"],"Median":r["med"],"Min":r["min"],"Max":r["max"]} for r in rows])
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine='openpyxl') as writer:
        result.to_excel(writer,index=False,sheet_name="Full Data")
        s2=s2df(stats_rows); g2=g2df(groupby_rows)
        if not s2.empty: s2.to_excel(writer,index=False,sheet_name="Overall Stats")
        if not g2.empty: g2.to_excel(writer,index=False,sheet_name="Category Stats")
        specs=[("date","Date Wise","1F4E79","GateOut Date"),("dow","Day of Week","375623","GateOut DayOfWeek"),
               ("week","Week Wise","833C00","GateOut WeekNo"),("month","Month Wise","7B2D8B","GateOut Month")]
        for key,sname,_,gcol in specs:
            rows=time_data.get(key,[])
            if rows: t2df(rows,gcol).to_excel(writer,index=False,sheet_name=sname)
        SC={"Full Data":"1F4E79","Overall Stats":"375623","Category Stats":"7B2D8B",
            "Date Wise":"1F4E79","Day of Week":"375623","Week Wise":"833C00","Month Wise":"7B2D8B"}
        for sn,ws in writer.sheets.items(): style_ws(ws,SC.get(sn,"1F4E79"))
        wf=writer.sheets["Full Data"]
        clm={col:get_column_letter(idx) for idx,col in enumerate(result.columns,1)}
        for cn in dt_col_names:
            if cn in clm:
                cl=clm[cn]
                for rn in range(2,len(result)+2):
                    cell=wf[f"{cl}{rn}"]
                    if isinstance(cell.value,str) and cell.value!="":
                        try: cell.value=pd.to_datetime(cell.value,dayfirst=True).to_pydatetime()
                        except: pass
                    if cell.value and not isinstance(cell.value,str): cell.number_format="DD-MM-YYYY HH:MM:SS"
        for cn in tat_cols_set:
            if cn in clm:
                ci=result.columns.tolist().index(cn)+1
                wf.cell(1,ci).fill=PatternFill("solid",start_color="375623")
                wf.cell(1,ci).font=Font(name="Arial",size=10,bold=True,color="FFFFFF")
                cl=get_column_letter(ci)
                for rn in range(2,len(result)+2):
                    cell=wf[f"{cl}{rn}"]
                    if isinstance(cell.value,str) and cell.value!="":
                        frac=hms_str_to_fraction(cell.value)
                        if frac is not None: cell.value=frac
                    if cell.value not in (None,""): cell.number_format="[HH]:MM:SS"
                    cell.fill=PatternFill("solid",start_color="E2EFDA")
                    cell.font=Font(name="Arial",size=9,bold=True,color="375623")
        TCS={"GateOut Date":("FCE4D6","833C00","DD-MM-YYYY"),"GateOut DayOfWeek":("E2EFDA","375623","@"),
             "GateOut WeekNo":("FFF2CC","7F6000","@"),"GateOut Month":("EAD1DC","7B2D8B","@")}
        for cn,(bg,fg,nf) in TCS.items():
            if cn in result.columns:
                ci=result.columns.tolist().index(cn)+1;cl=get_column_letter(ci)
                wf.cell(1,ci).fill=PatternFill("solid",start_color=fg)
                wf.cell(1,ci).font=Font(name="Arial",size=10,bold=True,color="FFFFFF")
                for rn in range(2,len(result)+2):
                    cell=wf[f"{cl}{rn}"];cell.fill=PatternFill("solid",start_color=bg)
                    cell.font=Font(name="Arial",size=9,bold=True,color=fg)
                    cell.number_format=nf;cell.alignment=Alignment(horizontal="center",vertical="center")
        for key,sname,_,gcol in specs:
            if sname in writer.sheets:
                rows=time_data.get(key,[])
                if rows: fmt_time_ws(writer.sheets[sname],len(rows)+1)
        if "Overall Stats" in writer.sheets and stats_rows: fmt_time_ws(writer.sheets["Overall Stats"],len(stats_rows)+1)
        if "Category Stats" in writer.sheets and groupby_rows: fmt_time_ws(writer.sheets["Category Stats"],len(groupby_rows)+1)
    buf.seek(0); return buf

# ─────────────────────────────────────────────────────────────
# HTML DASHBOARD
# ─────────────────────────────────────────────────────────────
DASHBOARD_HTML = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&family=Fira+Code:wght@400;500&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#060912;--s1:#0d1117;--s2:#131929;--s3:#1a2235;
  --p:#6366f1;--p2:#818cf8;--cyan:#22d3ee;--green:#10b981;--amber:#f59e0b;--red:#f43f5e;
  --text:#e2e8f0;--muted:#64748b;--dim:#334155;--border:rgba(99,102,241,0.15);
  --r:14px;--font:'Plus Jakarta Sans',sans-serif;--mono:'Fira Code',monospace;
}
html{scroll-behavior:smooth}
body{background:var(--bg);color:var(--text);font-family:var(--font);min-height:100vh;overflow-x:hidden}
body::before{
  content:'';position:fixed;inset:0;z-index:0;pointer-events:none;
  background:
    radial-gradient(ellipse 60% 40% at 10% 10%,rgba(99,102,241,0.07),transparent 60%),
    radial-gradient(ellipse 50% 50% at 90% 80%,rgba(34,211,238,0.05),transparent 60%);
}
/* NAV */
.nav{
  position:sticky;top:0;z-index:100;height:54px;
  background:rgba(6,9,18,0.88);backdrop-filter:blur(20px);
  border-bottom:1px solid var(--border);
  display:flex;align-items:center;padding:0 20px;gap:8px;
}
.nav-brand{font-weight:800;font-size:1rem;letter-spacing:-0.01em;
  background:linear-gradient(135deg,#a5b4fc,#22d3ee);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  margin-right:20px;white-space:nowrap;flex-shrink:0;
}
.nav-tabs{display:flex;gap:2px;flex:1;overflow-x:auto;scrollbar-width:none}
.nav-tabs::-webkit-scrollbar{display:none}
.nav-tab{
  border:none;border-radius:8px;padding:6px 14px;cursor:pointer;background:transparent;
  color:var(--muted);font-family:var(--font);font-size:0.78rem;font-weight:600;
  letter-spacing:0.03em;text-transform:uppercase;transition:all 0.2s;white-space:nowrap;
}
.nav-tab:hover{color:var(--text);background:rgba(99,102,241,0.07)}
.nav-tab.active{
  color:#fff;background:linear-gradient(135deg,rgba(99,102,241,0.22),rgba(129,140,248,0.14));
  border:1px solid rgba(99,102,241,0.28);
}
.nav-right{margin-left:auto;display:flex;align-items:center;gap:6px;flex-shrink:0}
.npill{font-size:0.68rem;font-weight:700;padding:3px 10px;border-radius:20px;letter-spacing:0.05em;text-transform:uppercase}
.npill-mode{background:rgba(99,102,241,0.12);border:1px solid rgba(99,102,241,0.25);color:var(--p2)}
.npill-live{background:rgba(16,185,129,0.08);border:1px solid rgba(16,185,129,0.25);color:var(--green);display:flex;align-items:center;gap:4px}
.npill-live::before{content:'';width:5px;height:5px;border-radius:50%;background:var(--green);box-shadow:0 0 5px var(--green);animation:blink 1.5s ease-in-out infinite}
@keyframes blink{0%,100%{opacity:1}50%{opacity:0.25}}
/* PAGES */
.page{display:none;animation:fadein 0.3s ease}
.page.active{display:block}
@keyframes fadein{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}
.content{padding:20px 24px;max-width:1380px;margin:0 auto}
/* SECTION */
.sec{font-size:0.62rem;font-weight:700;letter-spacing:0.2em;text-transform:uppercase;
  color:var(--p2);display:flex;align-items:center;gap:8px;margin:24px 0 14px}
.sec::after{content:'';flex:1;height:1px;background:linear-gradient(90deg,rgba(99,102,241,0.3),transparent)}
/* SUMMARY CHIPS */
.chips{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:4px}
.chip{background:var(--s2);border:1px solid var(--border);border-radius:10px;padding:10px 16px;
  transition:all 0.25s;cursor:default}
.chip:hover{border-color:rgba(99,102,241,0.35);transform:translateY(-2px);box-shadow:0 0 20px rgba(99,102,241,0.1)}
.chip-val{font-family:var(--mono);font-size:1.2rem;color:var(--cyan);line-height:1}
.chip-lbl{font-size:0.6rem;color:var(--muted);text-transform:uppercase;letter-spacing:0.08em;margin-top:3px}
.chip-val.ok{color:var(--green)}.chip-val.err{color:var(--red)}
/* PIPELINE */
.pipeline{background:var(--s2);border:1px solid var(--border);border-radius:var(--r);
  padding:18px 20px;position:relative;overflow:hidden;margin-bottom:8px}
.pipeline::before{content:'';position:absolute;left:0;top:0;bottom:0;width:3px;
  background:linear-gradient(180deg,var(--p),var(--cyan));border-radius:3px 0 0 3px}
.pl-hd{font-size:0.65rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:0.12em;margin-bottom:12px}
.pl-nodes{display:flex;align-items:center;flex-wrap:wrap;gap:6px}
.pl-node{background:rgba(99,102,241,0.07);border:1px solid rgba(99,102,241,0.2);
  border-radius:9px;padding:6px 12px;font-family:var(--mono);font-size:0.75rem;color:var(--p2);
  cursor:default;transition:all 0.22s cubic-bezier(0.34,1.56,0.64,1);position:relative}
.pl-node:hover{background:rgba(99,102,241,0.16);border-color:var(--p);color:#fff;
  transform:scale(1.06) translateY(-2px);box-shadow:0 0 16px rgba(99,102,241,0.28)}
.pl-node .tip{position:absolute;bottom:calc(100%+7px);left:50%;transform:translateX(-50%) scale(0.9) translateY(4px);
  background:var(--s1);border:1px solid var(--border);border-radius:7px;
  padding:5px 9px;font-size:0.68rem;color:var(--text);white-space:nowrap;
  opacity:0;pointer-events:none;transition:all 0.18s;font-family:var(--font);box-shadow:0 6px 18px rgba(0,0,0,0.5);z-index:10}
.pl-node:hover .tip{opacity:1;transform:translateX(-50%) scale(1) translateY(0)}
.pl-arr{color:rgba(99,102,241,0.3);font-size:1rem;flex-shrink:0}
.pl-stages{margin-top:10px;display:flex;flex-wrap:wrap;gap:5px}
.pl-pill{background:rgba(16,185,129,0.06);border:1px solid rgba(16,185,129,0.18);
  border-radius:20px;padding:2px 9px;font-family:var(--mono);font-size:0.68rem;color:var(--green);
  transition:all 0.2s;cursor:default}
.pl-pill:hover{background:rgba(16,185,129,0.14);box-shadow:0 0 8px rgba(16,185,129,0.2)}
/* KPI GRID */
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(155px,1fr));gap:10px}
.kpi{background:var(--s2);border:1px solid var(--border);border-radius:var(--r);
  padding:16px 14px;position:relative;overflow:hidden;
  transition:all 0.3s cubic-bezier(0.34,1.56,0.64,1);cursor:default}
.kpi::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;
  background:linear-gradient(90deg,var(--p),var(--cyan),var(--green));background-size:200% auto;
  animation:gflow 3s linear infinite}
@keyframes gflow{0%{background-position:0%}100%{background-position:200%}}
.kpi:hover{border-color:rgba(99,102,241,0.38);transform:translateY(-5px) scale(1.01);
  box-shadow:0 0 35px rgba(99,102,241,0.14),0 14px 35px rgba(0,0,0,0.4)}
.kpi-stage{font-size:0.62rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:0.14em;margin-bottom:8px}
.kpi-avg{font-family:var(--mono);font-size:1.7rem;color:var(--cyan);line-height:1;
  text-shadow:0 0 18px rgba(34,211,238,0.4)}
.kpi-albl{font-size:0.58rem;color:var(--dim);margin-bottom:8px}
.kpi-med{font-family:var(--mono);font-size:1.15rem;color:var(--green);
  text-shadow:0 0 14px rgba(16,185,129,0.4)}
.kpi-mlbl{font-size:0.58rem;color:var(--dim);margin-bottom:10px}
.kpi-bar{height:3px;background:rgba(255,255,255,0.05);border-radius:2px;overflow:hidden;margin-bottom:10px}
.kpi-bar-fill{height:100%;border-radius:2px;background:linear-gradient(90deg,var(--p),var(--cyan))}
.kpi-foot{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:4px}
.kf-lbl{font-size:0.55rem;color:var(--dim);text-transform:uppercase;letter-spacing:0.08em;font-weight:700}
.kf-val{font-family:var(--mono);font-size:0.68rem;color:var(--muted);margin-top:1px}
/* FILTER BAR */
.fbar{display:flex;align-items:center;gap:10px;flex-wrap:wrap;
  background:var(--s2);border:1px solid var(--border);border-radius:var(--r);
  padding:10px 16px;margin-bottom:14px}
.flbl{font-size:0.65rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:0.1em;white-space:nowrap}
.fsel{background:var(--s3);border:1px solid var(--border);border-radius:8px;
  padding:6px 10px;color:var(--text);font-family:var(--font);font-size:0.8rem;
  outline:none;cursor:pointer;transition:all 0.2s}
.fsel:hover,.fsel:focus{border-color:var(--p);box-shadow:0 0 0 2px rgba(99,102,241,0.12)}
.ftg{display:flex;gap:3px;background:var(--s3);border-radius:8px;padding:3px}
.ft{border:none;border-radius:6px;padding:5px 11px;cursor:pointer;
  font-family:var(--font);font-size:0.73rem;font-weight:600;
  color:var(--muted);background:transparent;transition:all 0.18s}
.ft.active{background:linear-gradient(135deg,var(--p),var(--p2));color:#fff;
  box-shadow:0 2px 10px rgba(99,102,241,0.3)}
.finput{background:var(--s3);border:1px solid var(--border);border-radius:8px;
  padding:6px 12px;color:var(--text);font-family:var(--font);font-size:0.8rem;
  outline:none;transition:all 0.2s;width:220px}
.finput:focus{border-color:var(--p);box-shadow:0 0 0 2px rgba(99,102,241,0.12)}
.finput::placeholder{color:var(--dim)}
/* CHART CARDS */
.cg{display:grid;gap:14px;margin-bottom:14px}
.cg2{grid-template-columns:1fr 1fr}.cg1{grid-template-columns:1fr}
@media(max-width:860px){.cg2{grid-template-columns:1fr}}
.ccard{background:var(--s2);border:1px solid var(--border);border-radius:var(--r);
  padding:18px;transition:border-color 0.3s}
.ccard:hover{border-color:rgba(99,102,241,0.28)}
.cc-t{font-size:0.8rem;font-weight:700;color:var(--text);margin-bottom:3px}
.cc-s{font-size:0.67rem;color:var(--muted);margin-bottom:14px}
.cwrap{position:relative;height:210px}
/* INNER TABS */
.itabs{display:flex;gap:3px;background:var(--s2);border-radius:10px;padding:4px;
  margin-bottom:14px;border:1px solid var(--border)}
.itab{flex:1;border:none;background:transparent;border-radius:7px;
  padding:7px 4px;font-family:var(--font);font-size:0.73rem;font-weight:600;
  color:var(--muted);cursor:pointer;transition:all 0.18s;text-align:center}
.itab:hover{color:var(--text)}
.itab.active{background:linear-gradient(135deg,var(--p),var(--p2));color:#fff;
  box-shadow:0 0 14px rgba(99,102,241,0.3)}
.ipanel{display:none}
.ipanel.active{display:block;animation:fadein 0.28s ease}
/* TABLE */
.twrap{overflow-x:auto;border-radius:12px;border:1px solid var(--border)}
table{width:100%;border-collapse:collapse;font-size:0.78rem}
thead{position:sticky;top:0;z-index:5}
th{background:var(--s2);color:var(--p2);font-size:0.63rem;font-weight:700;
  text-transform:uppercase;letter-spacing:0.12em;padding:9px 11px;
  border-bottom:1px solid var(--border);white-space:nowrap;text-align:left}
td{padding:7px 11px;color:var(--text);border-bottom:1px solid rgba(255,255,255,0.03);
  font-family:var(--mono);font-size:0.76rem}
tr:hover td{background:rgba(99,102,241,0.05)}
tr:last-child td{border-bottom:none}
.td-c{color:var(--cyan);font-weight:500}.td-g{color:var(--green);font-weight:500}
.td-badge{display:inline-block;background:rgba(99,102,241,0.1);border:1px solid rgba(99,102,241,0.2);
  border-radius:5px;padding:1px 7px;font-size:0.7rem;color:var(--p2)}
.td-bw{display:flex;align-items:center;gap:7px}
.td-bt{flex:1;height:4px;background:rgba(255,255,255,0.05);border-radius:2px;overflow:hidden;min-width:50px}
.td-bf{height:100%;border-radius:2px;background:linear-gradient(90deg,var(--p),var(--cyan))}
/* TIME CARDS */
.tcg{display:grid;grid-template-columns:repeat(auto-fill,minmax(105px,1fr));gap:8px;margin-top:10px}
.tc{background:var(--s2);border:1px solid var(--border);border-radius:11px;
  padding:10px 7px;text-align:center;transition:all 0.28s cubic-bezier(0.34,1.56,0.64,1);cursor:default}
.tc:hover{border-color:rgba(99,102,241,0.4);transform:translateY(-4px) scale(1.04);
  box-shadow:0 0 22px rgba(99,102,241,0.14)}
.tc-hd{background:linear-gradient(135deg,var(--p),var(--p2));border-radius:6px;
  padding:4px 5px;margin-bottom:7px;font-size:0.67rem;font-weight:700;color:#fff}
.tc-av{font-family:var(--mono);font-size:0.9rem;color:var(--cyan)}
.tc-al{font-size:0.52rem;color:var(--dim);text-transform:uppercase;letter-spacing:0.07em;margin-bottom:5px}
.tc-mv{font-family:var(--mono);font-size:0.76rem;color:var(--green)}
.tc-ml{font-size:0.52rem;color:var(--dim);text-transform:uppercase;letter-spacing:0.07em;margin-bottom:5px}
.tc-tr{background:rgba(255,255,255,0.04);border-radius:5px;padding:3px;font-size:0.7rem;color:var(--muted);font-family:var(--mono)}
/* CHATBOT */
#fabw{position:fixed;bottom:26px;right:26px;z-index:9999;font-family:var(--font)}
.fabc{position:relative;width:56px;height:56px}
.fabc::before,.fabc::after{content:'';position:absolute;inset:-8px;border-radius:50%;
  border:1.5px solid rgba(99,102,241,0.28);animation:pr 2.8s ease-out infinite}
.fabc::after{animation-delay:1.4s}
@keyframes pr{0%{transform:scale(0.82);opacity:0.7}100%{transform:scale(1.5);opacity:0}}
.fab{width:56px;height:56px;border-radius:50%;border:none;cursor:pointer;
  background:linear-gradient(135deg,#6366f1,#8b5cf6,#22d3ee);
  display:flex;align-items:center;justify-content:center;font-size:22px;
  position:absolute;inset:0;z-index:2;
  box-shadow:0 0 0 1px rgba(99,102,241,0.4),0 0 28px rgba(99,102,241,0.5),0 8px 22px rgba(0,0,0,0.5);
  animation:bob 3.5s ease-in-out infinite;transition:all 0.3s cubic-bezier(0.34,1.56,0.64,1);overflow:hidden}
.fab::after{content:'';position:absolute;inset:0;border-radius:50%;
  background:linear-gradient(135deg,rgba(255,255,255,0.14),transparent 55%)}
.fab:hover{transform:scale(1.14) rotate(8deg)!important;
  box-shadow:0 0 0 1px rgba(99,102,241,0.5),0 0 48px rgba(99,102,241,0.7),0 12px 32px rgba(0,0,0,0.6)!important}
@keyframes bob{0%,100%{transform:translateY(0)}50%{transform:translateY(-7px)}}
.fabt{position:absolute;right:66px;top:50%;
  transform:translateY(-50%) translateX(5px) scale(0.93);
  background:var(--s2);border:1px solid rgba(99,102,241,0.22);
  border-radius:9px;padding:6px 12px;font-size:0.74rem;color:var(--muted);
  white-space:nowrap;opacity:0;pointer-events:none;
  transition:all 0.22s cubic-bezier(0.34,1.56,0.64,1);box-shadow:0 5px 18px rgba(0,0,0,0.5)}
.fabt::after{content:'';position:absolute;right:-5px;top:50%;transform:translateY(-50%) rotate(45deg);
  width:8px;height:8px;background:var(--s2);
  border-top:1px solid rgba(99,102,241,0.22);border-right:1px solid rgba(99,102,241,0.22)}
#fabw:hover .fabt{opacity:1;transform:translateY(-50%) translateX(0) scale(1)}
.cwin{position:fixed;bottom:94px;right:26px;z-index:9998;width:340px;
  border-radius:18px;overflow:hidden;background:var(--s1);
  border:1px solid rgba(99,102,241,0.22);
  box-shadow:0 0 55px rgba(99,102,241,0.1),0 22px 55px rgba(0,0,0,0.65);
  display:none;flex-direction:column;animation:ci 0.36s cubic-bezier(0.34,1.56,0.64,1)}
.cwin.open{display:flex}
@keyframes ci{from{opacity:0;transform:scale(0.8) translateY(18px);transform-origin:bottom right}to{opacity:1;transform:scale(1) translateY(0)}}
.chd{padding:13px 15px;display:flex;align-items:center;gap:10px;
  background:linear-gradient(135deg,#6366f1,#8b5cf6);position:relative;overflow:hidden}
.chd::before{content:'';position:absolute;inset:0;background:linear-gradient(135deg,rgba(255,255,255,0.1),transparent)}
.cav{width:34px;height:34px;border-radius:50%;flex-shrink:0;background:rgba(255,255,255,0.18);
  border:2px solid rgba(255,255,255,0.28);display:flex;align-items:center;justify-content:center;
  font-size:16px;z-index:1;animation:avp 2.5s ease-in-out infinite}
@keyframes avp{0%,100%{box-shadow:0 0 0 0 rgba(255,255,255,0.28)}50%{box-shadow:0 0 0 6px rgba(255,255,255,0)}}
.chi{z-index:1}
.chn{font-weight:700;font-size:0.86rem;color:#fff}
.chs{font-size:0.65rem;color:rgba(255,255,255,0.7);display:flex;align-items:center;gap:3px;margin-top:1px}
.chs::before{content:'';width:5px;height:5px;border-radius:50%;background:#34d399;box-shadow:0 0 5px #34d399;animation:blink 1.5s ease-in-out infinite}
.cxb{margin-left:auto;background:rgba(255,255,255,0.13);border:1px solid rgba(255,255,255,0.18);
  color:#fff;width:26px;height:26px;border-radius:50%;cursor:pointer;font-size:12px;
  display:flex;align-items:center;justify-content:center;transition:all 0.2s;z-index:1;flex-shrink:0}
.cxb:hover{background:rgba(255,255,255,0.28);transform:rotate(90deg)}
.cms{flex:1;height:245px;overflow-y:auto;padding:12px;display:flex;flex-direction:column;gap:7px;
  scrollbar-width:thin;scrollbar-color:rgba(99,102,241,0.15) transparent}
.msg{max-width:88%;padding:8px 11px;border-radius:12px;font-size:0.78rem;line-height:1.5;
  animation:mp 0.28s cubic-bezier(0.34,1.56,0.64,1)}
@keyframes mp{from{opacity:0;transform:scale(0.92) translateY(7px)}to{opacity:1;transform:scale(1) translateY(0)}}
.msg.bot{background:rgba(99,102,241,0.08);border:1px solid rgba(99,102,241,0.14);
  color:var(--text);align-self:flex-start;border-bottom-left-radius:3px}
.msg.user{background:linear-gradient(135deg,rgba(99,102,241,0.18),rgba(139,92,246,0.12));
  border:1px solid rgba(99,102,241,0.22);color:var(--text);align-self:flex-end;border-bottom-right-radius:3px}
.typi{align-self:flex-start;background:rgba(99,102,241,0.08);border:1px solid rgba(99,102,241,0.14);
  border-radius:12px;border-bottom-left-radius:3px;padding:9px 13px;display:none}
.dots{display:flex;gap:3px}
.db{width:5px;height:5px;border-radius:50%;background:var(--p2);animation:bd 1.2s ease-in-out infinite}
.db:nth-child(2){animation-delay:0.2s}.db:nth-child(3){animation-delay:0.4s}
@keyframes bd{0%,100%{transform:translateY(0);opacity:0.4}50%{transform:translateY(-5px);opacity:1}}
.qb-row{padding:7px 11px;display:flex;gap:4px;flex-wrap:wrap;border-top:1px solid rgba(99,102,241,0.07)}
.qb{background:rgba(99,102,241,0.06);border:1px solid rgba(99,102,241,0.14);
  border-radius:20px;padding:3px 8px;cursor:pointer;font-size:0.67rem;color:var(--muted);
  transition:all 0.18s;font-family:var(--font)}
.qb:hover{background:rgba(99,102,241,0.15);color:var(--p2);border-color:rgba(99,102,241,0.32);transform:scale(1.04)}
.cir{padding:9px 11px;border-top:1px solid rgba(99,102,241,0.07);display:flex;gap:6px;align-items:center}
.ci{flex:1;background:rgba(99,102,241,0.04);border:1px solid rgba(99,102,241,0.16);
  border-radius:9px;padding:7px 11px;color:var(--text);font-family:var(--font);
  font-size:0.78rem;outline:none;transition:all 0.2s}
.ci:focus{border-color:var(--p);box-shadow:0 0 0 2px rgba(99,102,241,0.12);background:rgba(99,102,241,0.08)}
.ci::placeholder{color:var(--dim)}
.csb{width:34px;height:34px;border-radius:8px;border:none;flex-shrink:0;
  background:linear-gradient(135deg,var(--p),var(--p2));color:#fff;cursor:pointer;
  font-size:13px;display:flex;align-items:center;justify-content:center;
  transition:all 0.22s cubic-bezier(0.34,1.56,0.64,1)}
.csb:hover{transform:scale(1.12) rotate(5deg);box-shadow:0 0 14px rgba(99,102,241,0.5)}
.empty{text-align:center;padding:2.5rem 1.5rem;color:var(--dim);font-size:0.85rem}
*{scrollbar-width:thin;scrollbar-color:rgba(99,102,241,0.15) transparent}
*::-webkit-scrollbar{width:4px;height:4px}
*::-webkit-scrollbar-thumb{background:rgba(99,102,241,0.18);border-radius:2px}
</style>
</head>
<body>

<nav class="nav">
  <div class="nav-brand">⚡ TAT Intelligence</div>
  <div class="nav-tabs">
    <button class="nav-tab active" onclick="go('overview',this)">📊 Overview</button>
    <button class="nav-tab" onclick="go('charts',this)">📈 Charts</button>
    <button class="nav-tab" onclick="go('timedim',this)">📅 Time</button>
    <button class="nav-tab" onclick="go('category',this)">🏷 Category</button>
    <button class="nav-tab" onclick="go('data',this)">🗂 Data</button>
  </div>
  <div class="nav-right">
    <span class="npill npill-mode" id="modeP"></span>
    <span class="npill npill-live">LIVE</span>
  </div>
</nav>

<!-- OVERVIEW -->
<div class="page active content" id="page-overview">
  <div class="sec">Summary</div>
  <div class="chips" id="chips"></div>
  <div class="sec">Process Pipeline</div>
  <div class="pipeline" id="pipeline"></div>
  <div class="sec">TAT Stage Performance</div>
  <div class="kpi-grid" id="kpig"></div>
</div>

<!-- CHARTS -->
<div class="page content" id="page-charts">
  <div class="sec">Charts</div>
  <div class="fbar">
    <span class="flbl">Stage</span>
    <select class="fsel" id="cStage" onchange="renderCharts()"></select>
    <span class="flbl" style="margin-left:12px">Metric</span>
    <div class="ftg">
      <button class="ft active" onclick="setCM('avg',this)">Average</button>
      <button class="ft" onclick="setCM('med',this)">Median</button>
    </div>
  </div>
  <div class="cg cg2" id="cgrid"></div>
</div>

<!-- TIME DIM -->
<div class="page content" id="page-timedim">
  <div class="sec">Time Dimension Analysis</div>
  <div class="itabs">
    <button class="itab active" onclick="setTD('date',this)">📆 Date</button>
    <button class="itab" onclick="setTD('dow',this)">📅 Day of Week</button>
    <button class="itab" onclick="setTD('week',this)">🗓 Week</button>
    <button class="itab" onclick="setTD('month',this)">🗃 Month</button>
  </div>
  <div class="fbar">
    <span class="flbl">Stage</span>
    <select class="fsel" id="tdStage" onchange="renderTD()"></select>
    <span class="flbl" style="margin-left:12px">Metric</span>
    <div class="ftg">
      <button class="ft active" onclick="setTM('avg',this)">Average</button>
      <button class="ft" onclick="setTM('med',this)">Median</button>
      <button class="ft" onclick="setTM('both',this)">Both</button>
    </div>
  </div>
  <div id="tdcont"></div>
</div>

<!-- CATEGORY -->
<div class="page content" id="page-category">
  <div class="sec" id="catSec">Category Breakdown</div>
  <div id="catcont"></div>
</div>

<!-- DATA -->
<div class="page content" id="page-data">
  <div class="sec">Full Data <span id="dcnt" style="font-weight:400;text-transform:none;letter-spacing:0;font-size:0.75rem;color:var(--muted)"></span></div>
  <div class="fbar">
    <span class="flbl">Search</span>
    <input class="finput" id="tsearch" placeholder="Filter rows…" oninput="filterTbl()">
  </div>
  <div class="twrap" style="max-height:500px;overflow-y:auto" id="dtbl"></div>
</div>

<!-- CHATBOT -->
<div id="fabw">
  <div class="fabc"><button class="fab" onclick="togChat()">🤖</button></div>
  <div class="fabt">💡 TAT Assistant</div>
</div>
<div class="cwin" id="cwin">
  <div class="chd">
    <div class="cav">🤖</div>
    <div class="chi">
      <div class="chn">TAT Assistant</div>
      <div class="chs">Online · always ready</div>
    </div>
    <button class="cxb" onclick="togChat()">✕</button>
  </div>
  <div class="cms" id="cms">
    <div class="msg bot">👋 Hi! I'm your <strong>TAT Assistant</strong>. Ask about any stage, bottleneck, or metric!</div>
    <div class="msg bot">💡 Try: <em>"What is GI-GO?"</em> · <em>"How to reduce TAT?"</em> · <em>"Find bottleneck"</em></div>
  </div>
  <div class="typi" id="typi"><div class="dots"><span class="db"></span><span class="db"></span><span class="db"></span></div></div>
  <div class="qb-row">
    <button class="qb" onclick="qa('What is GI-GO?')">GI-GO?</button>
    <button class="qb" onclick="qa('What is GW-TW?')">GW-TW?</button>
    <button class="qb" onclick="qa('Reduce TAT')">Reduce TAT</button>
    <button class="qb" onclick="qa('Find bottleneck')">Bottleneck</button>
    <button class="qb" onclick="qa('Average vs Median')">Avg vs Med</button>
  </div>
  <div class="cir">
    <input class="ci" id="cinp" placeholder="Ask a question…" onkeydown="if(event.key==='Enter')sendChat()">
    <button class="csb" onclick="sendChat()">➤</button>
  </div>
</div>

<script>
const D=__DATA__;
const PV=__PREVIEW__;
let cMetric='avg',tdMetric='avg',activeTD='date';
const charts={};
const COLORS=['rgba(99,102,241,0.75)','rgba(34,211,238,0.75)','rgba(16,185,129,0.75)','rgba(245,158,11,0.75)','rgba(244,63,94,0.75)','rgba(139,92,246,0.75)'];
const CBORDER=['#6366f1','#22d3ee','#10b981','#f59e0b','#f43f5e','#8b5cf6'];
const PTIPS={YardIn:'Vehicle enters yard',GateIn:'Gate registration',GrossWeight:'Gross weight measured',TareWeight:'Tare weight measured',GateOut:'Vehicle exits',LoadingIn:'Loading begins',LoadingOut:'Loading complete'};
const PL={inbound:{nodes:['YardIn','GateIn','GrossWeight','TareWeight','GateOut'],stages:['YI-GI','GI-GW','GW-TW','TW-GO','GI-GO']},outbound:{nodes:['GrossWeight','LoadingIn','LoadingOut','TareWeight','GateOut'],stages:['GW-LI','LI-LO','LO-TW','TW-GO','GW-GO']}};

document.addEventListener('DOMContentLoaded',()=>{
  document.getElementById('modeP').textContent=D.mode==='inbound'?'📥 Inbound':'📤 Outbound';
  buildChips();buildPipeline();buildKPIs();populateFilters();
  renderCharts();renderTD();renderCategory();renderData();
});

function go(id,btn){
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.nav-tab').forEach(b=>b.classList.remove('active'));
  document.getElementById('page-'+id).classList.add('active');btn.classList.add('active');
  if(id==='charts')setTimeout(()=>renderCharts(),60);
}

function buildChips(){
  const match=D.total===D.output;
  document.getElementById('chips').innerHTML=`
    <div class="chip"><div class="chip-val">${D.total.toLocaleString()}</div><div class="chip-lbl">Uploaded Rows</div></div>
    <div class="chip"><div class="chip-val">${D.output.toLocaleString()}</div><div class="chip-lbl">Output Rows</div></div>
    <div class="chip"><div class="chip-val ${match?'ok':'err'}">${match?'✅ Match':'❌ Mismatch'}</div><div class="chip-lbl">Row Count</div></div>
    <div class="chip"><div class="chip-val">${D.stats.length}</div><div class="chip-lbl">TAT Stages</div></div>
    ${D.sel_cat&&D.sel_cat!=='-- None --'?`<div class="chip"><div class="chip-val" style="font-size:0.85rem;color:var(--p2)">${D.sel_cat}</div><div class="chip-lbl">Group-by</div></div>`:''}
  `;
}

function buildPipeline(){
  const p=PL[D.mode];
  const nodes=p.nodes.map((n,i)=>`${i>0?'<span class="pl-arr">→</span>':''}<div class="pl-node">${n}<div class="tip">${PTIPS[n]||n}</div></div>`).join('');
  const pills=p.stages.map(s=>`<div class="pl-pill">${s}</div>`).join('');
  document.getElementById('pipeline').innerHTML=`<div class="pl-hd">Process Flow</div><div class="pl-nodes">${nodes}</div><div class="pl-stages">${pills}</div>`;
}

function buildKPIs(){
  const mx=Math.max(...D.stats.map(s=>s.avg_min||0));
  document.getElementById('kpig').innerHTML=D.stats.map(s=>{
    const pct=mx>0?Math.round((s.avg_min/mx)*100):0;
    return `<div class="kpi">
      <div class="kpi-stage">${s.stage}</div>
      <div class="kpi-avg">${s.avg}</div><div class="kpi-albl">Average</div>
      <div class="kpi-med">${s.med}</div><div class="kpi-mlbl">Median</div>
      <div class="kpi-bar"><div class="kpi-bar-fill" style="width:${pct}%"></div></div>
      <div class="kpi-foot">
        <div><div class="kf-lbl">Min</div><div class="kf-val">${s.min}</div></div>
        <div><div class="kf-lbl">Max</div><div class="kf-val">${s.max}</div></div>
        <div><div class="kf-lbl">Valid</div><div class="kf-val">${s.valid}</div></div>
        <div><div class="kf-lbl">Blank</div><div class="kf-val">${s.blank}</div></div>
      </div>
    </div>`;
  }).join('');
}

function populateFilters(){
  const stages=D.stats.map(s=>s.stage);
  const opts=`<option value="all">All Stages</option>`+stages.map(s=>`<option value="${s}">${s}</option>`).join('');
  document.getElementById('cStage').innerHTML=opts;
  document.getElementById('tdStage').innerHTML=opts;
}

function setCM(m,btn){cMetric=m;document.querySelectorAll('#page-charts .ft').forEach(b=>b.classList.remove('active'));btn.classList.add('active');renderCharts();}
function setTM(m,btn){tdMetric=m;document.querySelectorAll('#page-timedim .ft').forEach(b=>b.classList.remove('active'));btn.classList.add('active');renderTD();}
function setTD(dim,btn){activeTD=dim;document.querySelectorAll('.itab').forEach(b=>b.classList.remove('active'));btn.classList.add('active');renderTD();}

function dc(id){if(charts[id]){charts[id].destroy();delete charts[id];}}
function baseOpts(yLabel,horiz=false){
  return{responsive:true,maintainAspectRatio:false,indexAxis:horiz?'y':'x',
    plugins:{legend:{labels:{color:'#94a3b8',font:{family:'Fira Code',size:10}}},
      tooltip:{callbacks:{label:ctx=>`${ctx.dataset.label}: ${(horiz?ctx.parsed.x:ctx.parsed.y)?.toFixed?.(1)??'N/A'} min`}}},
    scales:{
      x:{ticks:{color:'#64748b',font:{size:9}},grid:{color:'rgba(255,255,255,0.04)'}},
      y:{ticks:{color:'#64748b',font:{size:9}},grid:{color:'rgba(255,255,255,0.04)'}}
    }};
}

function renderCharts(){
  const sf=document.getElementById('cStage').value;
  let stages=D.stats.filter(s=>sf==='all'||s.stage===sf);
  const g=document.getElementById('cgrid'); g.innerHTML='';
  if(!stages.length){g.innerHTML='<div class="empty">No data.</div>';return;}

  // Bar
  g.insertAdjacentHTML('beforeend',`<div class="ccard"><div class="cc-t">Stage Comparison</div><div class="cc-s">${cMetric==='avg'?'Average':'Median'} TAT (minutes)</div><div class="cwrap"><canvas id="cbar"></canvas></div></div>`);
  dc('cbar');charts['cbar']=new Chart(document.getElementById('cbar'),{type:'bar',
    data:{labels:stages.map(s=>s.stage),datasets:[{label:cMetric==='avg'?'Average':'Median',
      data:stages.map(s=>cMetric==='avg'?s.avg_min:s.med_min),
      backgroundColor:stages.map((_,i)=>COLORS[i%COLORS.length]),
      borderColor:stages.map((_,i)=>CBORDER[i%CBORDER.length]),
      borderWidth:1,borderRadius:6,borderSkipped:false}]},
    options:baseOpts('Minutes')});

  // Range
  g.insertAdjacentHTML('beforeend',`<div class="ccard"><div class="cc-t">Min / Max Range</div><div class="cc-s">Best & worst case per stage (minutes)</div><div class="cwrap"><canvas id="crange"></canvas></div></div>`);
  dc('crange');charts['crange']=new Chart(document.getElementById('crange'),{type:'bar',
    data:{labels:stages.map(s=>s.stage),datasets:[
      {label:'Min',data:stages.map(s=>s.min_min),backgroundColor:'rgba(16,185,129,0.65)',borderColor:'#10b981',borderWidth:1,borderRadius:4},
      {label:'Max',data:stages.map(s=>s.max_min),backgroundColor:'rgba(244,63,94,0.65)',borderColor:'#f43f5e',borderWidth:1,borderRadius:4}
    ]},options:{...baseOpts('Minutes'),plugins:{...baseOpts('Minutes').plugins,tooltip:{callbacks:{label:ctx=>`${ctx.dataset.label}: ${ctx.parsed.y?.toFixed?.(1)} min`}}}}});

  if(sf!=='all'){
    const s=stages[0];
    g.insertAdjacentHTML('beforeend',`<div class="ccard"><div class="cc-t">Data Quality — ${s.stage}</div><div class="cc-s">Valid vs Blank rows</div><div class="cwrap"><canvas id="cdnt"></canvas></div></div>`);
    dc('cdnt');charts['cdnt']=new Chart(document.getElementById('cdnt'),{type:'doughnut',
      data:{labels:['Valid','Blank'],datasets:[{data:[s.valid,s.blank],
        backgroundColor:['rgba(16,185,129,0.75)','rgba(244,63,94,0.3)'],
        borderColor:['#10b981','#f43f5e'],borderWidth:1}]},
      options:{responsive:true,maintainAspectRatio:false,cutout:'65%',
        plugins:{legend:{labels:{color:'#94a3b8',font:{family:'Fira Code',size:11}}},
          tooltip:{callbacks:{label:ctx=>`${ctx.label}: ${ctx.parsed} rows`}}}}});

    g.insertAdjacentHTML('beforeend',`<div class="ccard"><div class="cc-t">Avg vs Median — ${s.stage}</div><div class="cc-s">Side-by-side comparison (minutes)</div><div class="cwrap"><canvas id="ccomp"></canvas></div></div>`);
    dc('ccomp');charts['ccomp']=new Chart(document.getElementById('ccomp'),{type:'bar',
      data:{labels:['Average','Median','Min','Max'],datasets:[{label:s.stage,
        data:[s.avg_min,s.med_min,s.min_min,s.max_min],
        backgroundColor:['rgba(99,102,241,0.75)','rgba(16,185,129,0.75)','rgba(34,211,238,0.75)','rgba(245,158,11,0.75)'],
        borderColor:['#6366f1','#10b981','#22d3ee','#f59e0b'],borderWidth:1,borderRadius:6,borderSkipped:false}]},
      options:baseOpts('Minutes')});
  }
}

function renderTD(){
  const rows=D.time[activeTD]||[];
  const sf=document.getElementById('tdStage').value;
  const filtered=rows.filter(r=>sf==='all'||r.stage===sf);
  const cont=document.getElementById('tdcont');
  if(!filtered.length){cont.innerHTML='<div class="empty">No data for this dimension.</div>';return;}

  const stages=[...new Set(filtered.map(r=>r.stage))];
  const labels=[...new Set(filtered.map(r=>r.label))];

  // Line chart
  const chartHTML=`<div class="ccard" style="margin-bottom:14px"><div class="cc-t">TAT Trend — ${activeTD==='date'?'Date':activeTD==='dow'?'Day of Week':activeTD==='week'?'Week':'Month'}</div><div class="cwrap" style="height:230px"><canvas id="tdchart"></canvas></div></div>`;
  let tableHTML='';
  if(tdMetric==='both'){
    tableHTML=`<div class="twrap"><table><thead><tr><th>Period</th><th>Stage</th><th class="td-c">Average</th><th class="td-g">Median</th><th>Min</th><th>Max</th><th>Trips</th></tr></thead><tbody>`;
    filtered.forEach(r=>{tableHTML+=`<tr><td>${r.label}</td><td><span class="td-badge">${r.stage}</span></td><td class="td-c">${r.avg}</td><td class="td-g">${r.med}</td><td style="color:var(--green)">${r.min}</td><td style="color:var(--amber)">${r.max}</td><td style="color:var(--muted)">${r.count}</td></tr>`;});
    tableHTML+=`</tbody></table></div>`;
  }
  let cardsHTML='';
  if(sf!=='all'){
    cardsHTML=`<div class="tcg">`+filtered.map(r=>`
      <div class="tc"><div class="tc-hd">${r.label}</div>
      <div class="tc-av">${r.avg}</div><div class="tc-al">Avg</div>
      <div class="tc-mv">${r.med}</div><div class="tc-ml">Med</div>
      <div class="tc-tr">${r.count} trips</div></div>`).join('')+`</div>`;
  }
  cont.innerHTML=chartHTML+(tdMetric==='both'?tableHTML:cardsHTML);

  dc('tdchart');
  const mKey=tdMetric==='avg'?'avg_min':'med_min';
  const datasets=stages.map((st,i)=>({
    label:st,tension:0.4,pointRadius:4,pointHoverRadius:7,borderWidth:2,fill:true,
    borderColor:CBORDER[i%CBORDER.length],
    backgroundColor:COLORS[i%COLORS.length].replace('0.75','0.1'),
    data:labels.map(lbl=>{const r=filtered.find(r=>r.label===lbl&&r.stage===st);return r?(tdMetric==='avg'?r.avg_min:r.med_min):null;}),
  }));
  charts['tdchart']=new Chart(document.getElementById('tdchart'),{type:'line',
    data:{labels,datasets},
    options:{responsive:true,maintainAspectRatio:false,interaction:{mode:'index',intersect:false},
      plugins:{legend:{labels:{color:'#94a3b8',font:{family:'Fira Code',size:10}}},
        tooltip:{callbacks:{label:ctx=>`${ctx.dataset.label}: ${ctx.parsed.y?.toFixed?.(1)??'N/A'} min`}}},
      scales:{
        x:{ticks:{color:'#64748b',font:{size:9}},grid:{color:'rgba(255,255,255,0.04)'}},
        y:{ticks:{color:'#64748b',font:{size:9}},grid:{color:'rgba(255,255,255,0.04)'}}
      }}});
}

function renderCategory(){
  const el=document.getElementById('catcont');
  if(!D.groupby||!D.groupby.length){
    el.innerHTML='<div class="empty">No category group-by selected.<br><small style="color:var(--dim)">Re-run analysis with a Group-by Category.</small></div>';return;
  }
  document.getElementById('catSec').textContent=`Category Breakdown — ${D.sel_cat}`;
  const cats=[...new Set(D.groupby.map(r=>r.category))];
  const stages=[...new Set(D.groupby.map(r=>r.stage))];
  const maxA=Math.max(...D.groupby.map(r=>r.avg_min||0));

  let html=`<div class="ccard" style="margin-bottom:14px">
    <div class="cc-t">Average TAT by ${D.sel_cat}</div>
    <div class="cwrap" style="height:${Math.max(200,cats.length*28)}px"><canvas id="catc"></canvas></div>
  </div>`;
  html+=`<div class="twrap"><table><thead><tr><th>Category</th><th>Stage</th><th class="td-c">Average</th><th class="td-g">Median</th><th>Trips</th><th style="min-width:110px">Bar</th></tr></thead><tbody>`;
  D.groupby.forEach(r=>{
    const pct=maxA>0?Math.round((r.avg_min/maxA)*100):0;
    html+=`<tr><td>${r.category}</td><td><span class="td-badge">${r.stage}</span></td><td class="td-c">${r.avg}</td><td class="td-g">${r.med}</td><td style="color:var(--muted)">${r.count}</td>
    <td><div class="td-bw"><div class="td-bt"><div class="td-bf" style="width:${pct}%"></div></div><span style="font-size:0.65rem;color:var(--muted)">${pct}%</span></div></td></tr>`;
  });
  html+=`</tbody></table></div>`;
  el.innerHTML=html;
  dc('catc');
  const datasets=stages.map((st,i)=>({
    label:st,data:cats.map(cat=>{const r=D.groupby.find(r=>r.stage===st&&r.category===cat);return r?r.avg_min:null;}),
    backgroundColor:COLORS[i%COLORS.length],borderColor:CBORDER[i%CBORDER.length],borderWidth:1,borderRadius:4,
  }));
  charts['catc']=new Chart(document.getElementById('catc'),{type:'bar',
    data:{labels:cats,datasets},
    options:{responsive:true,maintainAspectRatio:false,indexAxis:'y',
      plugins:{legend:{labels:{color:'#94a3b8',font:{family:'Fira Code',size:10}}},
        tooltip:{callbacks:{label:ctx=>`${ctx.dataset.label}: ${ctx.parsed.x?.toFixed?.(1)} min`}}},
      scales:{
        x:{ticks:{color:'#64748b',font:{size:9}},grid:{color:'rgba(255,255,255,0.04)'}},
        y:{ticks:{color:'#64748b',font:{size:9}},grid:{display:false}}
      }}});
}

let allRows=[],allCols=[];
function renderData(){
  allCols=PV.columns;allRows=PV.data;
  document.getElementById('dcnt').textContent=`(top ${allRows.length} rows)`;
  buildTbl(allRows);
}
function buildTbl(rows){
  const TAT=new Set(D.tat_cols);
  let html=`<table><thead><tr>${allCols.map(c=>`<th>${c}</th>`).join('')}</tr></thead><tbody>`;
  rows.forEach(row=>{
    html+='<tr>'+allCols.map((col,i)=>{
      const v=row[i]??'';
      return `<td${TAT.has(col)?' class="td-c"':''}>${v}</td>`;
    }).join('')+'</tr>';
  });
  html+=`</tbody></table>`;
  document.getElementById('dtbl').innerHTML=html;
}
function filterTbl(){
  const q=document.getElementById('tsearch').value.toLowerCase();
  buildTbl(q?allRows.filter(r=>r.some(v=>(v+'').toLowerCase().includes(q))):allRows);
}

// CHATBOT
function togChat(){const w=document.getElementById('cwin');w.classList.toggle('open');if(w.classList.contains('open'))setTimeout(()=>document.getElementById('cinp').focus(),300);}
function qa(q){document.getElementById('cinp').value=q;sendChat();}
function sendChat(){
  const inp=document.getElementById('cinp');const q=inp.value.trim();if(!q)return;
  addMsg(q,'user');inp.value='';
  const ti=document.getElementById('typi');ti.style.display='block';scrollC();
  setTimeout(()=>{ti.style.display='none';addMsg(getAns(q),'bot');},580+Math.random()*350);
}
function addMsg(t,w){
  const box=document.getElementById('cms');
  const d=document.createElement('div');d.className='msg '+w;d.innerHTML=t;
  box.insertBefore(d,document.getElementById('typi'));scrollC();
}
function scrollC(){const m=document.getElementById('cms');m.scrollTop=m.scrollHeight;}
const KB={
  'yi-gi':'🔵 <strong>YI-GI</strong> = GateIn − YardIn<br>Yard-to-gate wait. High values → parking congestion.',
  'gi-gw':'🔵 <strong>GI-GW</strong> = GrossWeight − GateIn<br>Gate entry to gross weigh. Delays = weighbridge queue.',
  'gw-tw':'🟢 <strong>GW-TW</strong> = TareWeight − GrossWeight<br>Usually the <em>longest stage</em> — loading/unloading time.',
  'tw-go':'🔵 <strong>TW-GO</strong> = GateOut − TareWeight<br>Post-tare paperwork & exit queue. Target: under 30 min.',
  'gi-go':'⚡ <strong>GI-GO</strong> = GateOut − GateIn<br><em>Headline KPI</em> — total time inside the plant.',
  'gw-li':'🟣 <strong>GW-LI</strong> = LoadingIn − GrossWeight<br>Wait before loading starts. High = dock scheduling issue.',
  'li-lo':'🟣 <strong>LI-LO</strong> = LoadingOut − LoadingIn<br>Actual loading duration.',
  'lo-tw':'🟣 <strong>LO-TW</strong> = TareWeight − LoadingOut<br>Post-loading tare queue.',
  'gw-go':'⚡ <strong>GW-GO</strong> = GateOut − GrossWeight<br>Total outbound plant TAT.',
  'average':'📈 <strong>Average</strong> = sum ÷ count. Sensitive to delay spikes. Always pair with Median.',
  'median':'📉 <strong>Median</strong> = middle value. Robust to outliers. If Avg >> Median, you have extreme delay events.',
  'reduce':'🚀 <strong>Reduce TAT:</strong><br>1️⃣ Add weighbridge capacity<br>2️⃣ Pre-register vehicles<br>3️⃣ Stagger shift timings<br>4️⃣ Streamline paperwork<br>5️⃣ Improve dock scheduling',
  'bottleneck':'🔍 <strong>Finding Bottlenecks:</strong><br>On the <strong>Charts tab</strong> → look for the stage with <strong>highest Average</strong> AND <strong>widest Max-Min gap</strong>. That stage needs the most attention.',
  'inbound':'📥 <strong>Inbound:</strong> YardIn→GateIn→GrossWeight→TareWeight→GateOut<br>Stages: YI-GI, GI-GW, GW-TW, TW-GO, GI-GO',
  'outbound':'📤 <strong>Outbound:</strong> GrossWeight→LoadingIn→LoadingOut→TareWeight→GateOut<br>Stages: GW-LI, LI-LO, LO-TW, TW-GO, GW-GO',
  'shift':'🕐 Use the <strong>Category tab</strong> with Shift as group-by to compare morning vs night shift performance.',
  'chart':'📈 Check the <strong>Charts tab</strong> for bar charts, range analysis, and doughnut data-quality view!',
  'time':'📅 The <strong>Time tab</strong> shows interactive line charts by date, day of week, week, and month.',
};
function getAns(q){
  const ql=q.toLowerCase();
  for(const[k,v] of Object.entries(KB))if(ql.includes(k))return v;
  if(ql.includes('hello')||ql.includes('hi'))return '👋 Hello! Ask me about any TAT stage or metric.';
  if(ql.includes('help'))return 'I can explain stages (YI-GI, GW-TW…), find bottlenecks, compare Avg vs Median, and suggest improvements!';
  return '🤔 Try asking about a stage like <strong>GI-GO</strong> or <strong>GW-TW</strong>, or ask how to <strong>reduce TAT</strong> or find <strong>bottlenecks</strong>.';
}
</script>
</body>
</html>
"""

def render_dashboard():
    result      = st.session_state["tat_result"]
    stats       = st.session_state["tat_stats"]
    groupby     = st.session_state["tat_groupby"]
    time_data   = st.session_state["tat_time"]
    total_rows  = st.session_state["tat_total"]
    sel_cat     = st.session_state.get("tat_sel_cat","-- None --")
    mode        = st.session_state.get("tat_mode","inbound")

    payload = json.dumps({
        "stats":stats,"groupby":groupby,"time":time_data,
        "total":total_rows,"output":len(result),
        "mode":mode,"sel_cat":sel_cat,"tat_cols":list(st.session_state["tat_tat_set"])
    }, ensure_ascii=False)

    all_prev=["Trip ID","Vehicle Number","Transporter Name","Shift","GateOut Date",
              "GateOut DayOfWeek","GateOut WeekNo","GateOut Month","YardIn","GateIn",
              "GrossWeight","TareWeight","GateOut","YI-GI","GI-GW","GW-TW","TW-GO",
              "GI-GO","GW-LI","LI-LO","LO-TW","GW-GO"]
    prev_cols=[c for c in all_prev if c in result.columns]
    preview_json = result[prev_cols].head(200).to_json(orient="split", default_handler=str)

    html = DASHBOARD_HTML.replace("__DATA__", payload).replace("__PREVIEW__", preview_json)
    components.html(html, height=820, scrolling=False)

# ─────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("<h2 style='color:#a5b4fc;font-size:1.1rem;margin-bottom:4px'>⚡ TAT Intelligence</h2>", unsafe_allow_html=True)
    st.markdown("<p style='color:#475569;font-size:0.75rem;margin-bottom:12px'>Turnaround Time Analytics</p>", unsafe_allow_html=True)
    st.divider()

    mode_sel = st.radio("Mode", ["📥 Inbound","📤 Outbound"],
                        index=0 if st.session_state.get("tat_mode","inbound")=="inbound" else 1)
    new_mode = "inbound" if "Inbound" in mode_sel else "outbound"
    if new_mode != st.session_state.get("tat_mode","inbound"):
        st.session_state["tat_mode"] = new_mode
        for k in ["tat_result","tat_stats","tat_groupby","tat_time","tat_dt_cols","tat_tat_set","tat_total","tat_gateout_col"]:
            st.session_state.pop(k, None)
        st.rerun()
    st.session_state["tat_mode"] = new_mode

    st.divider()
    uploaded = st.file_uploader(f"Upload {'Inbound' if new_mode=='inbound' else 'Outbound'} Excel",
                                type=["xlsx","xls"], key=f"up_{new_mode}")

    if uploaded:
        df = load_file(uploaded)
        total_rows = len(df)
        st.success(f"✅ {total_rows:,} rows · {len(df.columns)} cols")
        all_cols = ["-- Not Available --"] + df.columns.tolist()

        st.markdown("<div style='color:#64748b;font-size:0.72rem;text-transform:uppercase;letter-spacing:0.08em;margin:10px 0 4px'>Column Mapping</div>", unsafe_allow_html=True)

        if new_mode == "inbound":
            cy = st.selectbox("YardIn",     all_cols, index=auto_index(all_cols,"YardIn"),      key="yi")
            cg = st.selectbox("GateIn",     all_cols, index=auto_index(all_cols,"GateIn"),      key="gi")
            cw = st.selectbox("GrossWeight",all_cols, index=auto_index(all_cols,"GrossWeight"), key="gw")
            ct = st.selectbox("TareWeight", all_cols, index=auto_index(all_cols,"TareWeight"),  key="tw")
            co = st.selectbox("GateOut",    all_cols, index=auto_index(all_cols,"GateOut"),     key="go")
        else:
            cw  = st.selectbox("GrossWeight",   all_cols, index=auto_index(all_cols,"GrossWeight"),  key="gw")
            cli = st.selectbox("LoadingIn",      all_cols, index=auto_index(all_cols,"LoadingIn"),    key="li")
            clo = st.selectbox("LoadingOut",     all_cols, index=auto_index(all_cols,"LoadingOut"),   key="lo")
            ct  = st.selectbox("TareWeight",     all_cols, index=auto_index(all_cols,"TareWeight"),   key="tw")
            co  = st.selectbox("GateOut",        all_cols, index=auto_index(all_cols,"GateOut"),      key="go")
            cy  = st.selectbox("YardIn (opt)",   all_cols, index=auto_index(all_cols,"YardIn"),       key="yi")
            cg  = st.selectbox("GateIn (opt)",   all_cols, index=auto_index(all_cols,"GateIn"),       key="gi")

        st.markdown("<div style='color:#64748b;font-size:0.72rem;text-transform:uppercase;letter-spacing:0.08em;margin:10px 0 4px'>Options</div>", unsafe_allow_html=True)
        CAT_P = ["Transporter Name","Shift","Mat. Group","Material Group",
                 "Unloader Alias","Vehicle Number","Gate Entry Type","Supplier Name"]
        cat_f = ["-- None --"] + [c for c in CAT_P if c in df.columns]
        sel_cat = st.selectbox("Group-by Category", cat_f, key="cat")

        if st.button("⚡ Calculate TAT", use_container_width=True, type="primary"):
            result = df.copy()
            if new_mode == "inbound":
                dt_yi = to_dt(result[cy]) if cy!="-- Not Available --" else None
                dt_gi = to_dt(result[cg]) if cg!="-- Not Available --" else None
                dt_gw = to_dt(result[cw]) if cw!="-- Not Available --" else None
                dt_tw = to_dt(result[ct]) if ct!="-- Not Available --" else None
                dt_go = to_dt(result[co]) if co!="-- Not Available --" else None
                stages_list=[("YI-GI",dt_yi,dt_gi,"YardIn","GateIn"),("GI-GW",dt_gi,dt_gw,"GateIn","GrossWeight"),
                             ("GW-TW",dt_gw,dt_tw,"GrossWeight","TareWeight"),("TW-GO",dt_tw,dt_go,"TareWeight","GateOut"),
                             ("GI-GO",dt_gi,dt_go,"GateIn","GateOut")]
                used=[c for c in [cy,cg,cw,ct,co] if c!="-- Not Available --"]
                gateout=co
            else:
                dt_yi = to_dt(result[cy])  if cy !="-- Not Available --" else None
                dt_gi = to_dt(result[cg])  if cg !="-- Not Available --" else None
                dt_gw = to_dt(result[cw])  if cw !="-- Not Available --" else None
                dt_li = to_dt(result[cli]) if cli!="-- Not Available --" else None
                dt_lo = to_dt(result[clo]) if clo!="-- Not Available --" else None
                dt_tw = to_dt(result[ct])  if ct !="-- Not Available --" else None
                dt_go = to_dt(result[co])  if co !="-- Not Available --" else None
                stages_list=[]
                if dt_yi and dt_gi: stages_list.append(("YI-GI",dt_yi,dt_gi,"YardIn","GateIn"))
                if dt_gi and dt_gw: stages_list.append(("GI-GW",dt_gi,dt_gw,"GateIn","GrossWeight"))
                stages_list+=[("GW-LI",dt_gw,dt_li,"GrossWeight","LoadingIn"),
                              ("LI-LO",dt_li,dt_lo,"LoadingIn","LoadingOut"),
                              ("LO-TW",dt_lo,dt_tw,"LoadingOut","TareWeight"),
                              ("TW-GO",dt_tw,dt_go,"TareWeight","GateOut"),
                              ("GW-GO",dt_gw,dt_go,"GrossWeight","GateOut")]
                used=[c for c in [cy,cg,cw,cli,clo,ct,co] if c!="-- Not Available --"]
                gateout=co

            with st.spinner("Calculating TAT…"):
                result, tat_cols_set = calculate_stages(result, stages_list, st)
            save_state(result, tat_cols_set, used, sel_cat, total_rows, gateout)

        st.divider()
        if "tat_result" in st.session_state:
            buf = export_excel(
                st.session_state["tat_result"], st.session_state["tat_stats"],
                st.session_state["tat_groupby"], st.session_state["tat_time"],
                st.session_state["tat_dt_cols"], st.session_state["tat_tat_set"])
            fname = "Inbound_TAT.xlsx" if new_mode=="inbound" else "Outbound_TAT.xlsx"
            st.download_button("⬇️ Download Excel (7 sheets)", data=buf, file_name=fname,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
    else:
        st.info("👆 Upload an Excel file to begin.")
        for k in ["tat_result","tat_stats","tat_groupby","tat_time","tat_dt_cols","tat_tat_set","tat_total"]:
            st.session_state.pop(k, None)

# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
if "tat_result" in st.session_state:
    render_dashboard()
else:
    st.markdown("""
    <style>@keyframes fl{0%,100%{transform:translateY(0)}50%{transform:translateY(-14px)}}</style>
    <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;
         min-height:78vh;text-align:center;padding:2rem;font-family:'Plus Jakarta Sans',sans-serif">
      <div style="font-size:5rem;margin-bottom:1.5rem;animation:fl 3s ease-in-out infinite">⚡</div>
      <h1 style="font-size:2.4rem;font-weight:800;letter-spacing:-0.02em;margin-bottom:0.5rem;
          background:linear-gradient(135deg,#a5b4fc,#22d3ee);
          -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text">
        TAT Intelligence
      </h1>
      <p style="color:#475569;font-size:0.95rem;max-width:460px;line-height:1.7;margin-bottom:2rem">
        Upload your logistics data from the <strong style="color:#818cf8">sidebar</strong> to unlock
        interactive charts, time-dimension trends, category analysis, and an AI assistant.
      </p>
      <div style="display:flex;gap:10px;flex-wrap:wrap;justify-content:center">
        <div style="background:#0d1117;border:1px solid rgba(99,102,241,0.18);border-radius:10px;padding:12px 18px;font-size:0.78rem;color:#64748b">📊 Bar & Line Charts</div>
        <div style="background:#0d1117;border:1px solid rgba(99,102,241,0.18);border-radius:10px;padding:12px 18px;font-size:0.78rem;color:#64748b">📅 Time Dimension</div>
        <div style="background:#0d1117;border:1px solid rgba(99,102,241,0.18);border-radius:10px;padding:12px 18px;font-size:0.78rem;color:#64748b">🏷 Category Breakdown</div>
        <div style="background:#0d1117;border:1px solid rgba(99,102,241,0.18);border-radius:10px;padding:12px 18px;font-size:0.78rem;color:#64748b">🤖 AI Chat Assistant</div>
      </div>
    </div>
    """, unsafe_allow_html=True)