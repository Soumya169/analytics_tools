import streamlit as st
import pandas as pd
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from utils import load_file, to_dt, auto_index, sec_to_hms, min_to_hms
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Packer Analysis", page_icon="📦", layout="wide")

st.title("📦 Packer Analysis")
st.markdown("Analyse packer-wise performance — trip count, packing duration, material packed.")
st.markdown("---")

# ── INBOUND / OUTBOUND TOGGLE ────────────────────────────────
col_btn1, col_btn2, _ = st.columns([1, 1, 4])
with col_btn1:
    ib = st.button("📥 Inbound", use_container_width=True,
                   type="primary" if st.session_state.get("packer_mode","inbound") == "inbound" else "secondary")
with col_btn2:
    ob = st.button("📤 Outbound", use_container_width=True,
                   type="primary" if st.session_state.get("packer_mode","inbound") == "outbound" else "secondary")
if ib: st.session_state["packer_mode"] = "inbound"
if ob: st.session_state["packer_mode"] = "outbound"
mode = st.session_state.get("packer_mode", "inbound")

st.markdown(f"**Mode: {'📥 Inbound' if mode == 'inbound' else '📤 Outbound'}**")
st.markdown("---")

uploaded = st.file_uploader("📂 Upload Excel File", type=["xlsx","xls"], key="packer_upload")
if uploaded is None:
    st.info("👆 Upload your Excel file to begin Packer Analysis.")
    st.markdown("""
    **Expected columns for Packer Analysis:**
    - `Packer Name` — packer identifier
    - `Packing Start` / `Packing End` — packing start & end datetime
    - `Shift` — shift (A/B/C)
    - `Mat. Group` or `Material Group` — material category
    - `Challan Quantity` or `Net Weight` — quantity packed
    """)
    st.stop()

df = load_file(uploaded)
total_rows = len(df)
st.success(f"✅ Loaded: **{total_rows} rows, {len(df.columns)} columns**")

with st.expander("🔍 Debug Info"):
    st.write("**Columns:**", df.columns.tolist())
    st.dataframe(df.head(3), use_container_width=True)

st.markdown("---")
st.subheader("🔧 Map Your Columns")
all_cols = ["-- Not Available --"] + df.columns.tolist()

c1, c2, c3 = st.columns(3)
with c1:
    col_packer    = st.selectbox("Packer Name",              all_cols, index=auto_index(all_cols,"Packer Name"))
    col_pack_start= st.selectbox("Packing Start (datetime)", all_cols, index=auto_index(all_cols,"Packing Start"))
with c2:
    col_pack_end  = st.selectbox("Packing End (datetime)",   all_cols, index=auto_index(all_cols,"Packing End"))
    col_shift     = st.selectbox("Shift",                    all_cols, index=auto_index(all_cols,"Shift"))
with c3:
    col_matgrp    = st.selectbox("Material Group",           all_cols, index=auto_index(all_cols,"Mat. Group"))
    col_qty       = st.selectbox("Quantity / Net Weight",    all_cols, index=auto_index(all_cols,"Challan Quantity"))

st.markdown("---")

if st.button("📊 Run Packer Analysis", type="primary", use_container_width=True):

    result = df.copy()

    # Calculate Packing Duration
    if col_pack_start != "-- Not Available --" and col_pack_end != "-- Not Available --":
        dt_ps = to_dt(result[col_pack_start])
        dt_pe = to_dt(result[col_pack_end])
        if dt_ps is not None and dt_pe is not None:
            diff = (dt_pe - dt_ps).dt.total_seconds()
            result["Packing Duration"] = diff.apply(sec_to_hms)
            result["Pack_min"]         = diff / 60
            st.success(f"✅ Packing Duration calculated for {int((diff >= 0).sum())} rows")

    st.markdown("---")

    if col_packer != "-- Not Available --":
        st.subheader("📋 Packer-wise Summary")

        agg_dict = {"Trip Count": (col_packer, "count")}
        if "Pack_min" in result.columns:
            agg_dict["Avg Packing Duration (min)"] = ("Pack_min", "mean")
            agg_dict["Total Packing Time (min)"]   = ("Pack_min", "sum")
        if col_qty != "-- Not Available --":
            result[col_qty] = pd.to_numeric(result[col_qty], errors='coerce')
            agg_dict["Total Quantity"] = (col_qty, "sum")
            agg_dict["Avg Quantity"]   = (col_qty, "mean")

        summary = result.groupby(col_packer).agg(**{
            k: v for k, v in agg_dict.items()
        }).round(2).reset_index()

        if "Avg Packing Duration (min)" in summary.columns:
            summary["Avg Duration (HH:MM:SS)"] = summary["Avg Packing Duration (min)"].apply(min_to_hms)
            summary = summary.drop(columns=["Avg Packing Duration (min)"])

        st.dataframe(summary, use_container_width=True)

        # Shift-wise
        if col_shift != "-- Not Available --":
            st.markdown("---")
            st.subheader("🕐 Shift-wise Packer Performance")
            shift_grp = result.groupby([col_packer, col_shift]).size().reset_index(name="Trip Count")
            st.dataframe(shift_grp, use_container_width=True)

        # Material-wise
        if col_matgrp != "-- Not Available --":
            st.markdown("---")
            st.subheader("📦 Material-wise Packer Performance")
            mat_grp = result.groupby([col_packer, col_matgrp]).size().reset_index(name="Trip Count")
            if col_qty != "-- Not Available --":
                mat_qty = result.groupby([col_packer, col_matgrp])[col_qty].sum().round(2).reset_index()
                mat_grp = mat_grp.merge(mat_qty, on=[col_packer, col_matgrp])
                mat_grp.columns = [col_packer, col_matgrp, "Trip Count", "Total Quantity"]
            st.dataframe(mat_grp, use_container_width=True)
    else:
        st.warning("⚠️ Please map the Packer Name column to see summary.")

    # Full preview
    st.markdown("---")
    st.subheader(f"👁 Full Data — {len(result)} Rows")
    st.dataframe(result, use_container_width=True, height=400)

    # Download
    st.markdown("---")
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        result.to_excel(writer, index=False, sheet_name="Packer Data")
        if col_packer != "-- Not Available --":
            summary.to_excel(writer, index=False, sheet_name="Packer Summary")

        for sname in writer.sheets:
            ws = writer.sheets[sname]
            for cell in ws[1]:
                cell.fill = PatternFill("solid", start_color="7B2D8B")
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            ws.freeze_panes = "A2"

    buf.seek(0)
    st.download_button(
        f"⬇️ Download Packer Analysis Excel ({len(result)} rows)",
        data=buf,
        file_name="Packer_Analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, type="primary"
    )